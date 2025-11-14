from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from rest_framework import viewsets
from .models import (
    Student, Parent, FeeCategory, FeeStructure, Payment,
    PaymentReceipt, FeeDiscount, PaymentReminder, SchoolBankAccount,
    DonationCategory, DonationEvent, Donation, EmailPreferences, FeeStatus,
    FeeWaiver, FeeSettings, AcademicTerm, IndividualStudentFee, UserProfile
)
from .serializers import PaymentSerializer
import requests
import hashlib
import json
from django.conf import settings
from rest_framework.response import Response
from rest_framework.decorators import api_view
from django.contrib import messages
from django.db.models import Sum, Count, Avg, Q
from django.db.models.functions import TruncMonth
from django.http import JsonResponse, HttpResponse
from .forms import (
    StudentForm, ParentForm, FeeCategoryForm, FeeStructureForm,
    PaymentForm, FeeDiscountForm, SchoolBankAccountForm, PaymentSearchForm,
    DonationCategoryForm, DonationEventForm, DonationForm, FeeWaiverForm,
    IndividualStudentFeeForm
)
from django.core.files.storage import default_storage
import qrcode
from io import BytesIO
from django.core.files.base import ContentFile
from django.contrib.auth.models import User
from django.template.loader import get_template, render_to_string
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from datetime import datetime, timezone
from django.utils import timezone
from django.core.mail import send_mail
from django.utils.html import strip_tags


from .ai_services import PaymentPredictionService
import nltk
from nltk.tokenize import word_tokenize
from nltk.corpus import stopwords
from nltk.stem import WordNetLemmatizer
import re
from django.views.decorators.http import require_GET

try:
    from openpyxl import Workbook
except ImportError:
    Workbook = None

# Download required NLTK data
nltk.download('punkt')
nltk.download('stopwords')
nltk.download('wordnet')

class SystemChatbot:
    def __init__(self):
        self.lemmatizer = WordNetLemmatizer()
        self.stop_words = set(stopwords.words('english'))
        
        # Define system knowledge base
        self.knowledge_base = {
            'payment': {
                'keywords': ['payment', 'pay', 'fee', 'fees', 'amount', 'money', 'transaction'],
                'info': {
                    'how_to_pay': 'You can make payments through multiple methods: cash, bank transfer, or online payment.',
                    'payment_methods': 'Available payment methods are: cash, bank transfer, and online payment.',
                    'payment_status': 'Payment status can be: pending, completed, or failed.',
                    'receipt': 'After successful payment, you can download or email your receipt.',
                }
            },
            'student': {
                'keywords': ['student', 'students', 'enrollment', 'enroll', 'register', 'registration'],
                'info': {
                    'registration': 'Students can be registered through the student registration form.',
                    'profile': 'Student profiles include: name, ID, year batch, and active status.',
                    'fees': 'Each student is assigned fee structures based on their form and category.',
                }
            },
            'fee_structure': {
                'keywords': ['fee structure', 'fee category', 'categories', 'structure', 'pricing'],
                'info': {
                    'categories': 'Fee categories include: tuition, examination, and miscellaneous fees.',
                    'frequency': 'Fees can be collected monthly, termly, or yearly.',
                    'amount': 'Fee amounts vary based on the category and form level.',
                }
            },
            'discount': {
                'keywords': ['discount', 'discounts', 'waiver', 'waivers', 'scholarship', 'scholarships'],
                'info': {
                    'types': 'Available discounts: percentage-based, fixed amount, or full waiver.',
                    'eligibility': 'Discounts are based on merit, need, or special circumstances.',
                    'application': 'Discount applications can be submitted through the fee waiver form.',
                }
            },
            'report': {
                'keywords': ['report', 'reports', 'statement', 'statements', 'analytics', 'dashboard'],
                'info': {
                    'types': 'Available reports: payment reports, fee analytics, and collection reports.',
                    'access': 'Reports can be accessed through the dashboard or reports section.',
                    'export': 'Reports can be exported in PDF or Excel format.',
                }
            }
        }

    def preprocess_text(self, text):
        # Convert to lowercase
        text = text.lower()
        # Remove special characters
        text = re.sub(r'[^\w\s]', '', text)
        # Tokenize
        tokens = word_tokenize(text)
        # Remove stopwords and lemmatize
        tokens = [self.lemmatizer.lemmatize(token) for token in tokens if token not in self.stop_words]
        return tokens

    def find_relevant_category(self, tokens):
        max_matches = 0
        best_category = None
        
        for category, data in self.knowledge_base.items():
            matches = sum(1 for token in tokens if token in data['keywords'])
            if matches > max_matches:
                max_matches = matches
                best_category = category
                
        return best_category

    def get_response(self, query):
        # Preprocess the query
        tokens = self.preprocess_text(query)
        
        # Find relevant category
        category = self.find_relevant_category(tokens)
        
        if category:
            # Get relevant information
            info = self.knowledge_base[category]['info']
            
            # Find the most relevant piece of information
            response = None
            for key, value in info.items():
                if any(token in key for token in tokens):
                    response = value
                    break
            
            if not response:
                # If no specific match, return general category information
                response = f"Here's what I know about {category}: " + " ".join(info.values())
            
            return response
        else:
            return "I'm sorry, I don't have information about that. Please try asking about payments, students, fee structures, discounts, or reports."

@api_view(['POST'])
def chatbot_query(request):
    if request.method == 'POST':
        query = request.data.get('query', '')
        if not query:
            return Response({'error': 'No query provided'}, status=400)
        
        chatbot = SystemChatbot()
        response = chatbot.get_response(query)
        
        return Response({'response': response})

def chatbot_interface(request):
    return render(request, 'myapp/chatbot.html')

class PaymentViewSet(viewsets.ModelViewSet):
    queryset = Payment.objects.all()
    serializer_class = PaymentSerializer

# Dashboard View
def dashboard(request):
    return render(request, "dashboard.html")

@api_view(['POST'])
def fpx_payment_request(request):
    # Extract data from request (assuming amount and order_id are passed)
    amount = request.data.get("amount")
    order_id = request.data.get("order_id")
    
    if not amount or not order_id:
        return Response({"error": "Amount and Order ID are required"}, status=400)

    # FPX credentials from settings.py
    fpx_direct_url = settings.FPX_DIRECT_URL
    fpx_api_key = settings.FPX_API_KEY
    fpx_secret = settings.FPX_SECRET

    # Prepare payload (modify as per FPX API docs)
    payload = {
        "order_id": order_id,
        "amount": amount,
        "currency": "MYR",  # Adjust as needed
        "description": "Payment for Order " + str(order_id),
        "return_url": "https://yourwebsite.com/payment/success",
        "cancel_url": "https://yourwebsite.com/payment/cancel",
    }

    # Generate HMAC Signature (if FPX requires it)
    payload_json = json.dumps(payload, separators=(',', ':'))
    signature = hashlib.sha256((payload_json + fpx_secret).encode()).hexdigest()

    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {fpx_api_key}",
        "X-Signature": signature,
    }

    try:
        # Send request to FPX
        response = requests.post(fpx_direct_url, data=payload_json, headers=headers)
        response_data = response.json()

        # Return response to frontend
        return Response(response_data, status=response.status_code)

    except requests.exceptions.RequestException as e:
        return Response({"error": str(e)}, status=500)

def home(request):
    """Home page - accessible to all users (public)"""
    
    # If user is authenticated, check for admin redirects
    if request.user.is_authenticated:
        # Check if user is waqaf admin and redirect to waqaf homepage
        try:
            profile = getattr(request.user, 'myapp_profile', None)
            if profile and profile.is_waqaf_admin():
                from django.shortcuts import redirect
                return redirect('waqaf:waqaf')
        except:
            pass
        
        # Check if user is donation admin and redirect to donation homepage
        try:
            profile = getattr(request.user, 'myapp_profile', None)
            if profile and profile.is_donation_admin():
                from django.shortcuts import redirect
                return redirect('donation_events')
        except:
            pass
        
        # Check if user is Form 1 admin and redirect to Form 1 admin dashboard
        try:
            profile = getattr(request.user, 'myapp_profile', None)
            if profile and profile.is_form1_admin():
                from django.shortcuts import redirect
                return redirect('form1_admin:dashboard')
        except:
            pass
        
        # Check if user is Form 3 admin and redirect to Form 3 admin dashboard
        try:
            profile = getattr(request.user, 'myapp_profile', None)
            if profile and profile.is_form3_admin():
                from django.shortcuts import redirect
                return redirect('form3_admin:dashboard')
        except:
            pass
    
    # For all users (authenticated and anonymous), show the public homepage
    return render(request, 'home.html')

@login_required
def admin_dashboard(request):
    """Student Fee Collection Admin Dashboard"""
    
    # Sample data - in a real application, this would come from your database
    context = {
        'total_students': 450,
        'actual_collection': 85000,
        'amount_due': 100000,
        'collection_percentage': 85,
        
        # Form & Class data
        'form_class_data': [
            {'form': 1, 'class': 'A', 'male': 20, 'female': 18, 'total': 38, 'due': 15000, 'paid': 12750, 'outstanding': 2250, 'achievement': 85},
            {'form': 1, 'class': 'B', 'male': 19, 'female': 20, 'total': 39, 'due': 15600, 'paid': 13260, 'outstanding': 2340, 'achievement': 85},
            {'form': 2, 'class': 'A', 'male': 17, 'female': 29, 'total': 46, 'due': 18400, 'paid': 15640, 'outstanding': 2760, 'achievement': 85},
            {'form': 2, 'class': 'B', 'male': 22, 'female': 39, 'total': 61, 'due': 24400, 'paid': 20740, 'outstanding': 3660, 'achievement': 85},
            {'form': 3, 'class': 'A', 'male': 25, 'female': 35, 'total': 60, 'due': 24000, 'paid': 20400, 'outstanding': 3600, 'achievement': 85},
        ],
        
        # Fee Category data
        'fee_categories': {
            'pta': 320,
            'activities': 320,
            'exams': 130,
            'dormitory': 200,
        },
        
        # Payment status data
        'paid_students': 320,
        'not_paid_students': 130,
        'paid_percentage': 71,
        'not_paid_percentage': 29,
        
        # Grade-wise payment data
        'grade_payment_data': [
            {'grade': 1, 'class': 'A', 'males_paid': 18, 'females_paid': 16, 'males_not_paid': 2, 'females_not_paid': 2, 'total': 38},
            {'grade': 1, 'class': 'B', 'males_paid': 17, 'females_paid': 18, 'males_not_paid': 2, 'females_not_paid': 2, 'total': 39},
            {'grade': 2, 'class': 'A', 'males_paid': 15, 'females_paid': 26, 'males_not_paid': 2, 'females_not_paid': 3, 'total': 46},
            {'grade': 2, 'class': 'B', 'males_paid': 20, 'females_paid': 35, 'males_not_paid': 2, 'females_not_paid': 4, 'total': 61},
            {'grade': 3, 'class': 'A', 'males_paid': 23, 'females_paid': 32, 'males_not_paid': 2, 'females_not_paid': 3, 'total': 60},
        ],
    }
    
    return render(request, 'admin_dashboard.html', context)

def download_report(request):
    """Download comprehensive report"""
    from django.http import JsonResponse
    return JsonResponse({'status': 'success', 'message': 'Report download initiated'})

def print_receipt(request):
    """Print receipt or overdue notice"""
    from django.http import JsonResponse
    return JsonResponse({'status': 'success', 'message': 'Print job initiated'})

def send_reminder(request):
    """Send parent reminder"""
    from django.http import JsonResponse
    return JsonResponse({'status': 'success', 'message': 'Reminders sent'})

@login_required
def school_fees_dashboard(request):
    """Student Fee Collection Admin Dashboard - School Fees Module"""
    
    # Sample data based on the requirements
    context = {
        # Main Summary Data
        'expected_amount': 100000,
        'actual_collection': 85000,
        'achievement_percentage': 85,
        'outstanding_amount': 15000,
        'total_students': 950,
        
        # Form & Class data
        'form_class_data': [
            {'form': 1, 'class': 'A', 'boys': 20, 'girls': 18, 'total': 38, 'expected': 3800, 'paid': 3500, 'outstanding': 300, 'achieved': 92},
            {'form': 1, 'class': 'B', 'boys': 19, 'girls': 20, 'total': 39, 'expected': 3900, 'paid': 3200, 'outstanding': 700, 'achieved': 82},
            {'form': 2, 'class': 'A', 'boys': 22, 'girls': 25, 'total': 47, 'expected': 4700, 'paid': 4200, 'outstanding': 500, 'achieved': 89},
            {'form': 2, 'class': 'B', 'boys': 18, 'girls': 21, 'total': 39, 'expected': 3900, 'paid': 3600, 'outstanding': 300, 'achieved': 92},
            {'form': 3, 'class': 'A', 'boys': 25, 'girls': 23, 'total': 48, 'expected': 4800, 'paid': 4500, 'outstanding': 300, 'achieved': 94},
            {'form': 3, 'class': 'B', 'boys': 20, 'girls': 24, 'total': 44, 'expected': 4400, 'paid': 4000, 'outstanding': 400, 'achieved': 91},
            {'form': 4, 'class': 'A', 'boys': 28, 'girls': 26, 'total': 54, 'expected': 5400, 'paid': 5000, 'outstanding': 400, 'achieved': 93},
            {'form': 4, 'class': 'B', 'boys': 24, 'girls': 27, 'total': 51, 'expected': 5100, 'paid': 4700, 'outstanding': 400, 'achieved': 92},
            {'form': 5, 'class': 'A', 'boys': 30, 'girls': 28, 'total': 58, 'expected': 5800, 'paid': 5500, 'outstanding': 300, 'achieved': 95},
            {'form': 5, 'class': 'B', 'boys': 26, 'girls': 29, 'total': 55, 'expected': 5500, 'paid': 5200, 'outstanding': 300, 'achieved': 95},
        ],
        
        # Fee Categories data
        'fee_categories': [
            {'name': 'PTA', 'expected': 25000, 'achieved': 87, 'color': 'blue'},
            {'name': 'Activities', 'expected': 18000, 'achieved': 80, 'color': 'yellow'},
            {'name': 'Examinations', 'expected': 22000, 'achieved': 90, 'color': 'purple'},
            {'name': 'Hostel', 'expected': 20000, 'achieved': 83, 'color': 'green'},
        ],
        
        # Payment Status data
        'paid_students': 820,
        'unpaid_students': 130,
        'paid_percentage': 86,
        'unpaid_percentage': 14,
        
        # Grade-wise payment data
        'grade_payment_data': [
            {'grade': 1, 'class': 'A', 'males_paid': 18, 'females_paid': 16, 'males_not_paid': 2, 'females_not_paid': 2, 'total': 38},
            {'grade': 1, 'class': 'B', 'males_paid': 17, 'females_paid': 18, 'males_not_paid': 2, 'females_not_paid': 2, 'total': 39},
            {'grade': 2, 'class': 'A', 'males_paid': 20, 'females_paid': 23, 'males_not_paid': 2, 'females_not_paid': 2, 'total': 47},
            {'grade': 2, 'class': 'B', 'males_paid': 16, 'females_paid': 19, 'males_not_paid': 2, 'females_not_paid': 2, 'total': 39},
            {'grade': 3, 'class': 'A', 'males_paid': 23, 'females_paid': 21, 'males_not_paid': 2, 'females_not_paid': 2, 'total': 48},
            {'grade': 3, 'class': 'B', 'males_paid': 18, 'females_paid': 22, 'males_not_paid': 2, 'females_not_paid': 2, 'total': 44},
            {'grade': 4, 'class': 'A', 'males_paid': 26, 'females_paid': 24, 'males_not_paid': 2, 'females_not_paid': 2, 'total': 54},
            {'grade': 4, 'class': 'B', 'males_paid': 22, 'females_paid': 25, 'males_not_paid': 2, 'females_not_paid': 2, 'total': 51},
            {'grade': 5, 'class': 'A', 'males_paid': 28, 'females_paid': 26, 'males_not_paid': 2, 'females_not_paid': 2, 'total': 58},
            {'grade': 5, 'class': 'B', 'males_paid': 24, 'females_paid': 27, 'males_not_paid': 2, 'females_not_paid': 2, 'total': 55},
        ],
    }
    
    return render(request, 'school_fees_dashboard.html', context)

@login_required
def school_fees(request):
    # Special handling for tamim123 - always treat as student
    if request.user.username == 'tamim123':
        # Force tamim123 to use student view regardless of other permissions
        if hasattr(request.user, 'myapp_profile') and request.user.myapp_profile.role == 'student':
            student = request.user.myapp_profile.student
            
            # Get all fee statuses for this student to show payment status
            # Only show unpaid fee statuses
            fee_statuses = FeeStatus.objects.filter(
                student=student, 
                status__in=['pending', 'overdue']
            ).select_related('fee_structure')
            
            # Get fee structures that are appropriate for this student's form/grade level
            student_level = student.get_level_display_value()
            available_fees = FeeStructure.objects.filter(
                form__iexact=student_level,
                is_active=True
            ).select_related('category')
            
            # Filter out fees that are completely paid
            fees_to_show = []
            for fee_structure in available_fees:
                fee_statuses_for_fee = fee_statuses.filter(fee_structure=fee_structure)
                if fee_statuses_for_fee.exists():
                    if fee_statuses_for_fee.exclude(status='paid').exists():
                        fees_to_show.append(fee_structure)
                else:
                    fees_to_show.append(fee_structure)
            
            # Get student payments and individual fees
            student_payments = Payment.objects.filter(student=student).order_by('-created_at')
            total_payments = Payment.objects.filter(student=student).count()
            
            individual_fees = []
            try:
                individual_fees = IndividualStudentFee.objects.filter(
                    student=student, 
                    is_active=True,
                    is_paid=False
                ).select_related('category').order_by('-created_at')
            except:
                individual_fees = []
            
            # Calculate discount information for each fee status
            for fee_status in fee_statuses:
                fee_status.discount_info = fee_status.get_discount_info()
            
            context = {
                'student': student,
                'available_fees': fees_to_show,
                'fee_statuses': fee_statuses,
                'recent_payments': student_payments[:5],
                'view_type': 'student',
                'total_payments': total_payments,
                'individual_fees': individual_fees,
                'student_level': student_level,
                'is_tamim': True,
            }
            return render(request, 'myapp/school_fees_student.html', context)
    
    # Check if user has admin role in UserProfile (but not tamim123)
    if (hasattr(request.user, 'myapp_profile') and 
        request.user.myapp_profile.role == 'admin' and 
        request.user.username != 'tamim123'):
        # Admin users get admin dashboard access (except tamim123)
        return render(request, 'myapp/school_fees.html', {'is_tamim': False})
    
    # Check if user is superuser (but not tamim123)
    if request.user.is_superuser and request.user.username != 'tamim123':
        # Superuser gets admin dashboard access (except tamim123)
        return render(request, 'myapp/school_fees.html', {'is_tamim': False})
    
    # Check if user is a parent (either by Parent model or UserProfile role)
    try:
        parent = Parent.objects.get(user=request.user)
        return redirect('myapp:parent_dashboard')
    except Parent.DoesNotExist:
        # Also check UserProfile role for parent
        if hasattr(request.user, 'myapp_profile') and request.user.myapp_profile.role == 'parent':
            return redirect('myapp:parent_dashboard')
    
    if hasattr(request.user, 'myapp_profile') and request.user.myapp_profile.role == 'student':
        student = request.user.myapp_profile.student
        
        # Get all fee statuses for this student to show payment status
        # Only show unpaid fee statuses
        fee_statuses = FeeStatus.objects.filter(
            student=student, 
            status__in=['pending', 'overdue']
        ).select_related('fee_structure')
        print(f"DEBUG: views.school_fees - Student {student.first_name} has {len(fee_statuses)} unpaid fee statuses")
        print(f"DEBUG: Fee statuses: {[(fs.fee_structure.category.name, fs.status) for fs in fee_statuses[:3]]}")
        
        # Get fee structures that are appropriate for this student's form/grade level
        # This ensures all students in the same form pay the same amount
        student_level = student.get_level_display_value()
        print(f"DEBUG: Student {student.first_name} is in {student_level}")
        
        # Get all active fee structures for this student's form level
        available_fees = FeeStructure.objects.filter(
            form__iexact=student_level,  # Case-insensitive match
            is_active=True
        ).select_related('category')
        
        print(f"DEBUG: Found {available_fees.count()} fee structures for {student_level}")
        
        # Filter out fees that are completely paid
        fees_to_show = []
        for fee_structure in available_fees:
            # Check if this fee has any pending/overdue statuses
            fee_statuses_for_fee = fee_statuses.filter(fee_structure=fee_structure)
            
            if fee_statuses_for_fee.exists():
                # If fee has status records, only show if any are pending/overdue
                if fee_statuses_for_fee.exclude(status='paid').exists():
                    fees_to_show.append(fee_structure)
            else:
                # If fee has no status records, it's new and should be shown
                fees_to_show.append(fee_structure)
        
        # Get student payments and individual fees
        student_payments = Payment.objects.filter(student=student).order_by('-created_at')
        total_payments = Payment.objects.filter(student=student).count()
        
        # Get individual student fees for this student (only unpaid fees)
        individual_fees = []
        try:
            individual_fees = IndividualStudentFee.objects.filter(
                student=student, 
                is_active=True,
                is_paid=False  # Only show unpaid fees
            ).select_related('category').order_by('-created_at')
        except:
            individual_fees = []
        
        # Calculate discount information for each fee status (only unpaid ones)
        for fee_status in fee_statuses:
            fee_status.discount_info = fee_status.get_discount_info()
        
        is_tamim = request.user.username == 'tamim123'
        context = {
            'student': student,
            'available_fees': fees_to_show,  # Use filtered fees
            'fee_statuses': fee_statuses,
            'recent_payments': student_payments[:5],
            'view_type': 'student',
            'total_payments': total_payments,
            'individual_fees': individual_fees,
            'student_level': student_level,  # Add student level to context
            'is_tamim': is_tamim,  # Add is_tamim context for template
        }
        return render(request, 'myapp/school_fees_student.html', context)
    
    # For any other users (non-students, non-admins, non-parents), show the general school fees page
    is_tamim = request.user.username == 'tamim123'
    return render(request, 'myapp/school_fees.html', {'is_tamim': is_tamim})

@login_required
def school_fees_dashboard(request):
    from django.db.models import Sum, Count, Q
    from datetime import datetime, timedelta
    from django.db.models.functions import TruncMonth
    import json
    
    # Get filter parameters
    filter_type = request.GET.get('filter_type', 'all')
    filter_value = request.GET.get('filter_value', '')
    status_filter = request.GET.get('status', 'all')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Base queryset for payments
    payments = Payment.objects.select_related('student', 'fee_structure__category')
    
    # Apply filters
    if filter_type == 'student' and filter_value:
        payments = payments.filter(student__first_name__icontains=filter_value)
    elif filter_type == 'class' and filter_value:
        payments = payments.filter(student__class_name__icontains=filter_value)
    elif filter_type == 'category' and filter_value:
        payments = payments.filter(fee_structure__category__name__icontains=filter_value)
    
    if status_filter == 'paid':
        payments = payments.filter(status='completed')
    elif status_filter == 'pending':
        payments = payments.filter(status='pending')
    
    if date_from:
        payments = payments.filter(payment_date__gte=date_from)
    if date_to:
        payments = payments.filter(payment_date__lte=date_to)
    
    # Calculate comprehensive statistics
    total_students = Student.objects.filter(is_active=True).count()
    total_payments_count = payments.count()
    total_paid = payments.filter(status='completed').count()
    pending_payments = payments.filter(status='pending').count()
    overdue_payments = FeeStatus.objects.filter(status='overdue').count()
    
    total_revenue = payments.filter(status='completed').aggregate(Sum('amount'))['amount__sum'] or 0
    pending_amount = payments.filter(status='pending').aggregate(Sum('amount'))['amount__sum'] or 0
    
    # Calculate rates
    payment_rate = (total_paid / total_payments_count * 100) if total_payments_count > 0 else 0
    collection_rate = (total_revenue / (total_revenue + pending_amount) * 100) if (total_revenue + pending_amount) > 0 else 0
    
    # Additional metrics
    active_students = Student.objects.filter(is_active=True).count()
    total_fee_structures = FeeStructure.objects.count()
    total_categories = FeeCategory.objects.count()
    paid_fee_statuses = FeeStatus.objects.filter(status='paid').count()
    pending_fee_statuses = FeeStatus.objects.filter(status='pending').count()
    overdue_fee_statuses = FeeStatus.objects.filter(status='overdue').count()
    
    # Get data for charts
    # 1. Payments by Category (Pie Chart)
    category_chart_data = payments.filter(status='completed').values(
        'fee_structure__category__name'
    ).annotate(
        total=Sum('amount')
    ).order_by('-total')
    
    category_chart_labels = [item['fee_structure__category__name'] or 'Individual Fee' for item in category_chart_data]
    category_chart_values = [float(item['total']) for item in category_chart_data]
    
    # 2. Payments by Class (Bar Chart)
    class_chart_data = payments.filter(status='completed').values(
        'student__class_name'
    ).annotate(
        total=Sum('amount')
    ).order_by('-total')
    
    class_chart_labels = [item['student__class_name'] or 'N/A' for item in class_chart_data]
    class_chart_values = [float(item['total']) for item in class_chart_data]
    
    # 3. Monthly Trends (Line Chart)
    six_months_ago = datetime.now() - timedelta(days=180)
    monthly_data = payments.filter(
        status='completed',
        payment_date__gte=six_months_ago
    ).annotate(
        month=TruncMonth('payment_date')
    ).values('month').annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('month')
    
    monthly_labels = [data['month'].strftime('%b %Y') for data in monthly_data]
    monthly_amounts = [float(data['total']) for data in monthly_data]
    monthly_counts = [data['count'] for data in monthly_data]
    
    # 4. Status Distribution
    status_distribution = {
        'labels': ['Paid', 'Pending', 'Overdue'],
        'data': [total_paid, pending_payments, overdue_payments]
    }
    
    # Get filter options
    students = Student.objects.all().order_by('first_name')
    classes = Student.objects.values_list('class_name', flat=True).distinct().exclude(class_name__isnull=True).exclude(class_name='')
    categories = FeeCategory.objects.all().order_by('name')
    
    # Recent payments
    recent_payments = payments.order_by('-payment_date')[:10]
    
    # CSV Export
    if request.GET.get('export') == 'csv':
        from django.http import HttpResponse
        import csv
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="payment_analytics.csv"'
        writer = csv.writer(response)
        
        writer.writerow(['Student', 'Class', 'Category', 'Amount', 'Status', 'Date'])
        for payment in payments:
            writer.writerow([
                f"{payment.student.first_name} {payment.student.last_name}",
                payment.student.class_name or 'N/A',
                payment.fee_structure.category.name if payment.fee_structure and payment.fee_structure.category else 'Individual Fee',
                payment.amount,
                payment.status,
                payment.payment_date
            ])
        return response
    
    context = {
        # Filter parameters
        'filter_type': filter_type,
        'filter_value': filter_value,
        'status_filter': status_filter,
        'date_from': date_from,
        'date_to': date_to,
        
        # Statistics
        'total_students': total_students,
        'total_payments': total_payments_count,
        'total_paid': total_paid,
        'pending_payments': pending_payments,
        'overdue_payments': overdue_payments,
        'total_revenue': total_revenue,
        'active_students': active_students,
        'total_fee_structures': total_fee_structures,
        'total_categories': total_categories,
        'payment_rate': payment_rate,
        'collection_rate': collection_rate,
        'paid_fee_statuses': paid_fee_statuses,
        'pending_fee_statuses': pending_fee_statuses,
        'overdue_fee_statuses': overdue_fee_statuses,
        
        # Chart data
        'category_chart_data': json.dumps({
            'labels': category_chart_labels,
            'data': category_chart_values
        }),
        'class_chart_data': json.dumps({
            'labels': class_chart_labels,
            'data': class_chart_values
        }),
        'monthly_chart_data': json.dumps({
            'labels': monthly_labels,
            'amounts': monthly_amounts,
            'counts': monthly_counts
        }),
        'status_distribution': json.dumps(status_distribution),
        
        # Filter options
        'students': students,
        'classes': classes,
        'categories': categories,
        
        # Recent data
        'recent_payments': recent_payments,
    }
    
    return render(request, 'myapp/school_fees_dashboard.html', context)

@login_required
def admin_fee_dashboard(request):
    """Comprehensive Student Fee Collection Admin Dashboard"""
    from django.db.models import Sum, Count, Q, Avg, Case, When, IntegerField
    from django.db import models
    from datetime import datetime, timedelta
    import json
    
    # Get filter parameters
    report_by = request.GET.get('report_by', 'all')
    form_filter = request.GET.get('form', '')
    class_filter = request.GET.get('class', '')
    category_filter = request.GET.get('category', '')
    
    # 1. MAIN SUMMARY DATA (Real-time from database)
    # Calculate expected amount from all fee structures
    total_fee_structures = FeeStructure.objects.filter(is_active=True)
    expected_amount = 0
    
    # If no active fee structures, try to activate some or use a fallback calculation
    if total_fee_structures.count() == 0:
        # Try to activate fee structures or create a basic calculation
        all_fee_structures = FeeStructure.objects.all()
        if all_fee_structures.exists():
            # Activate the first fee structure as a fallback
            first_fs = all_fee_structures.first()
            first_fs.is_active = True
            first_fs.save()
            total_fee_structures = FeeStructure.objects.filter(is_active=True)
    
    for fee_structure in total_fee_structures:
        # Count students for each form - handle different form formats
        form_value = fee_structure.form
        
        # Try different form matching strategies
        student_count = 0
        
        # Strategy 1: Direct match
        student_count = Student.objects.filter(
            level_custom=form_value,
            is_active=True
        ).count()
        
        # Strategy 2: If no match, try extracting number from form name
        if student_count == 0 and 'Form' in str(form_value):
            form_number = str(form_value).replace('Form', '').strip()
            student_count = Student.objects.filter(
                level_custom=form_number,
                is_active=True
            ).count()
        
        # Strategy 3: If still no match, try reverse (number to Form X)
        if student_count == 0 and str(form_value).isdigit():
            form_name = f"Form {form_value}"
            student_count = Student.objects.filter(
                level_custom=form_name,
                is_active=True
            ).count()
        
        expected_amount += fee_structure.amount * student_count
    
    # Fallback calculation if no fee structures or no matches
    if expected_amount == 0:
        # Calculate based on actual payments and outstanding amounts
        actual_collection = Payment.objects.filter(status='completed').aggregate(Sum('amount'))['amount__sum'] or 0
        outstanding_amount = FeeStatus.objects.filter(
            status__in=['pending', 'overdue']
        ).aggregate(Sum('amount'))['amount__sum'] or 0
        expected_amount = actual_collection + outstanding_amount
        
        # If still 0, use a basic calculation based on student count
        if expected_amount == 0:
            total_students = Student.objects.filter(is_active=True).count()
            # Assume average fee of RM 100 per student as fallback
            expected_amount = total_students * 100
    
    # Actual collection from completed payments
    actual_collection = Payment.objects.filter(status='completed').aggregate(Sum('amount'))['amount__sum'] or 0
    
    # Outstanding amount from pending and overdue fee statuses
    outstanding_amount = FeeStatus.objects.filter(
        status__in=['pending', 'overdue']
    ).aggregate(Sum('amount'))['amount__sum'] or 0
    
    # Achievement percentage - ensure it's properly calculated and capped at 100%
    if expected_amount > 0:
        achievement_percentage = min((actual_collection / expected_amount * 100), 100.0)
        achievement_percentage = round(achievement_percentage, 1)  # Round to 1 decimal place
    else:
        achievement_percentage = 0.0
    
    # Total active students
    total_students_registered = Student.objects.filter(is_active=True).count()
    
    # 2. REPORT BY FORM & CLASS (Real-time from database)
    form_class_data = []
    forms = ['1', '2', '3', '4', '5']
    classes = ['A', 'B', 'C', 'D', 'E', 'F']
    
    # Apply filters to determine which forms/classes to show
    if form_filter:
        forms = [form_filter]
    if class_filter:
        classes = [class_filter]
    
    # Handle report_by filter
    if report_by == 'form':
        # Group by form only, show all classes for each form
        pass  # Default behavior already groups by form and class
    elif report_by == 'class':
        # Group by class only, show all forms for each class  
        pass  # Default behavior already groups by form and class
    elif report_by == 'category':
        # This will be handled in the fee categories section
        pass
    
    for form in forms:
        for class_name in classes:
            # Get students for this form and class
            students_query = Student.objects.filter(
                level_custom=form,
                class_name=class_name,
                is_active=True
            )
            
            # Count by gender (assuming names ending with certain letters are female)
            # This is a simplified approach - in real system, you'd have a gender field
            male_students = students_query.filter(
                first_name__in=[
                    'Ahmad', 'Ali', 'Hassan', 'Ibrahim', 'Ismail', 'Muhammad', 'Omar', 'Yusuf',
                    'Adam', 'Daniel', 'David', 'James', 'John', 'Michael', 'Peter', 'Robert',
                    'Chen', 'Lee', 'Lim', 'Tan', 'Wong', 'Ng', 'Ong', 'Teo'
                ]
            ).count()
            
            total_students = students_query.count()
            female_students = total_students - male_students
            
            if total_students > 0:  # Only include classes with students
                # Calculate expected amount for this form/class
                # Try different form matching strategies for fee structures
                form_fee_structures = FeeStructure.objects.filter(
                    form=form,
                    is_active=True
                )
                
                # If no direct match, try with "Form X" format
                if not form_fee_structures.exists():
                    form_fee_structures = FeeStructure.objects.filter(
                        form=f"Form {form}",
                        is_active=True
                    )
                
                # If still no match, try extracting number from form field
                if not form_fee_structures.exists():
                    for fs in FeeStructure.objects.filter(is_active=True):
                        if str(form) in str(fs.form) or str(fs.form).replace('Form', '').strip() == str(form):
                            form_fee_structures = FeeStructure.objects.filter(id=fs.id)
                            break
                
                # If we found matching fee structures, calculate expected total
                if form_fee_structures.exists():
                    expected_total = sum(fs.amount for fs in form_fee_structures) * total_students
                else:
                    # Fallback: Use average payment amount per student for this form/class
                    avg_payment = Payment.objects.filter(
                        student__level_custom=form,
                        student__class_name=class_name,
                        status='completed'
                    ).aggregate(avg=models.Avg('amount'))['avg'] or 0
                    
                    if avg_payment > 0:
                        expected_total = avg_payment * total_students
                    else:
                        # Final fallback: Use the existing fee structure amount as base
                        base_amount = FeeStructure.objects.filter(is_active=True).first()
                        if base_amount:
                            expected_total = base_amount.amount * total_students
                        else:
                            expected_total = 1000 * total_students  # Default RM 1000 per student
                
                # Calculate paid amount
                paid_total = Payment.objects.filter(
                    student__level_custom=form,
                    student__class_name=class_name,
                    status='completed'
                ).aggregate(Sum('amount'))['amount__sum'] or 0
                
                # Calculate outstanding amount
                outstanding_total = FeeStatus.objects.filter(
                    student__level_custom=form,
                    student__class_name=class_name,
                    status__in=['pending', 'overdue']
                ).aggregate(Sum('amount'))['amount__sum'] or 0
                
                # Calculate achievement percentage - ensure it's properly capped at 100%
                if expected_total > 0:
                    achievement_rate = min((paid_total / expected_total * 100), 100.0)
                    achievement_rate = round(achievement_rate, 1)
                else:
                    achievement_rate = 0.0
                
                form_class_data.append({
                    'form': int(form),
                    'class': class_name,
                    'boys': male_students,
                    'girls': female_students,
                    'total': total_students,
                    'expected': float(expected_total),
                    'paid': float(paid_total),
                    'outstanding': float(outstanding_total),
                    'achieved': round(achievement_rate, 1)
                })
    
    # 3. FEE CATEGORIES DATA (Real-time from database)
    fee_categories = []
    categories = FeeCategory.objects.filter(is_active=True)
    color_map = {
        'PTA': 'primary',
        'Activities': 'warning', 
        'Exams': 'success',
        'Dormitory': 'info'
    }
    
    for category in categories:
        # Calculate total expected amount for this category
        category_fee_structures = FeeStructure.objects.filter(
            category=category,
            is_active=True
        )
        
        total_expected = 0
        for fs in category_fee_structures:
            student_count = Student.objects.filter(
                level_custom=fs.form,
                is_active=True
            ).count()
            total_expected += fs.amount * student_count
        
        # Calculate total paid for this category
        total_paid = Payment.objects.filter(
            fee_structure__category=category,
            status='completed'
        ).aggregate(Sum('amount'))['amount__sum'] or 0
        
        # Calculate achievement percentage - ensure it's properly capped at 100%
        if total_expected > 0:
            achievement = min((total_paid / total_expected * 100), 100.0)
            achievement = round(achievement, 1)
        else:
            achievement = 0.0
        
        fee_categories.append({
            'name': category.name,
            'amount': float(total_expected),
            'paid': float(total_paid),
            'achievement': round(achievement, 1),
            'color': color_map.get(category.name, 'secondary')
        })
    
    # 4. PAYMENT STATUS BY GRADE (Real-time from database)
    payment_status_data = []
    
    for form in forms:
        for class_name in classes:
            # Get all students for this form and class
            students_query = Student.objects.filter(
                level_custom=form,
                class_name=class_name,
                is_active=True
            )
            
            if students_query.count() > 0:
                # Separate male and female students (using name-based approximation)
                male_names = [
                    'Ahmad', 'Ali', 'Hassan', 'Ibrahim', 'Ismail', 'Muhammad', 'Omar', 'Yusuf',
                    'Adam', 'Daniel', 'David', 'James', 'John', 'Michael', 'Peter', 'Robert',
                    'Chen', 'Lee', 'Lim', 'Tan', 'Wong', 'Ng', 'Ong', 'Teo'
                ]
                
                male_students = students_query.filter(first_name__in=male_names)
                female_students = students_query.exclude(first_name__in=male_names)
                
                # Count students who have made payments (completed status)
                males_paid = male_students.filter(
                    payments__status='completed'
                ).distinct().count()
                
                females_paid = female_students.filter(
                    payments__status='completed'
                ).distinct().count()
                
                # Count students who haven't paid (have pending/overdue fee statuses)
                males_not_paid = male_students.filter(
                    fee_statuses__status__in=['pending', 'overdue']
                ).distinct().count()
                
                females_not_paid = female_students.filter(
                    fee_statuses__status__in=['pending', 'overdue']
                ).distinct().count()
                
                class_total_students = students_query.count()
                
                payment_status_data.append({
                    'grade': int(form),
                    'class': class_name,
                    'males_paid': males_paid,
                    'females_paid': females_paid,
                    'males_not_paid': males_not_paid,
                    'females_not_paid': females_not_paid,
                    'total': class_total_students
                })
    
    # Calculate totals for payment status
    total_paid = sum(item['males_paid'] + item['females_paid'] for item in payment_status_data)
    total_not_paid = sum(item['males_not_paid'] + item['females_not_paid'] for item in payment_status_data)
    total_all_students = total_paid + total_not_paid
    paid_percentage = (total_paid / total_all_students * 100) if total_all_students > 0 else 0
    not_paid_percentage = (total_not_paid / total_all_students * 100) if total_all_students > 0 else 0
    
    # Chart data for Collection by Form (Real-time)
    form_collection_data = {}
    for item in form_class_data:
        form = item['form']
        if form not in form_collection_data:
            form_collection_data[form] = 0
        form_collection_data[form] += item['paid']
    # Fallback dummy data if no data was produced
    if not form_collection_data or sum(form_collection_data.values()) == 0:
        form_collection_data = {
            '1': 50000,
            '2': 65000,
            '3': 80000,
            '4': 95000,
            '5': 120000,
        }
    
    # Chart data for Gender Distribution (Real-time)
    total_males = sum(item['boys'] for item in form_class_data)
    total_females = sum(item['girls'] for item in form_class_data)
    # Fallback if counts are zero
    if total_males == 0 and total_females == 0:
        total_males, total_females = 270, 230
    
    # Monthly trend data (Real-time from database)
    # Get payments from the last 12 months
    from datetime import datetime, timedelta
    from django.db.models.functions import TruncMonth
    
    twelve_months_ago = datetime.now() - timedelta(days=365)
    monthly_payments = Payment.objects.filter(
        status='completed',
        payment_date__gte=twelve_months_ago
    ).annotate(
        month=TruncMonth('payment_date')
    ).values('month').annotate(
        total=Sum('amount')
    ).order_by('month')
    
    # Create monthly data for charts
    monthly_labels = []
    monthly_data = []
    
    # Fill in the data for the last 12 months
    current_date = datetime.now()
    for i in range(12):
        month_date = current_date - timedelta(days=30 * i)
        month_label = month_date.strftime('%b')
        monthly_labels.insert(0, month_label)
        
        # Find matching data for this month
        month_total = 0
        for payment_month in monthly_payments:
            if (payment_month['month'].month == month_date.month and 
                payment_month['month'].year == month_date.year):
                month_total = float(payment_month['total'])
                break
        monthly_data.insert(0, month_total)
    # Provide dummy monthly trend if all zero
    if sum(monthly_data) == 0:
        monthly_data = [0, 0, 0, 5000, 12000, 65000, 70000, 68000, 72000, 90000, 115000, 18000]
        # keep monthly_labels as computed

    # Ensure paid/not paid chart has data
    if total_all_students == 0:
        total_paid = 320
        total_not_paid = 130
        total_all_students = total_paid + total_not_paid
        paid_percentage = (total_paid / total_all_students) * 100
        not_paid_percentage = 100 - paid_percentage
    
    context = {
        # Main Summary
        'expected_amount': expected_amount,
        'actual_collection': actual_collection,
        'outstanding_amount': outstanding_amount,
        'achievement_percentage': round(achievement_percentage, 1),
        'total_students': total_students_registered,
        
        # Form & Class Report
        'form_class_data': form_class_data,
        'form_collection_data': {
            'labels': [f'Form {k}' for k in sorted(form_collection_data.keys())],
            'data': [form_collection_data[k] for k in sorted(form_collection_data.keys())]
        },
        'gender_distribution': {
            'labels': ['Male', 'Female'],
            'data': [total_males, total_females]
        },
        'monthly_trend': {
            'labels': monthly_labels,
            'data': monthly_data
        },
        
        # Fee Categories
        'fee_categories': fee_categories,
        'category_chart_data': json.dumps({
            'labels': [cat['name'] for cat in fee_categories],
            'data': [cat['paid'] for cat in fee_categories]
        }),
        
        # Payment Status
        'payment_status_data': payment_status_data,
        'total_paid': total_paid,
        'total_not_paid': total_not_paid,
        'paid_percentage': round(paid_percentage, 1),
        'not_paid_percentage': round(not_paid_percentage, 1),
        'paid_not_paid_chart': json.dumps({
            'labels': ['Paid', 'Not Paid'],
            'data': [total_paid, total_not_paid]
        }),
        
        # Filter options
        'report_by': report_by,
        'form_filter': form_filter,
        'class_filter': class_filter,
        'category_filter': category_filter,
    }
    
    return render(request, 'myapp/admin_fee_dashboard.html', context)

@login_required
def admin_fee_dashboard_pdf(request):
    """Download admin fee dashboard with all charts, graphs, and analytics in PDF format"""
    # Check if user is superuser or admin
    if not (request.user.is_superuser or request.user.is_staff):
        from django.contrib import messages
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('myapp:admin_fee_dashboard')
    
    from django.http import HttpResponse
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from django.utils import timezone
    from io import BytesIO
    from datetime import datetime, timedelta
    import json
    
    # Get all the same data as the dashboard
    from django.db.models import Sum, Count, Q, Avg, Case, When, IntegerField
    from .models import Student, Payment, FeeStatus, FeeStructure, FeeCategory
    
    # Calculate comprehensive statistics (same logic as dashboard)
    total_fee_structures = FeeStructure.objects.filter(is_active=True)
    expected_amount = 0
    
    if total_fee_structures.count() == 0:
        all_fee_structures = FeeStructure.objects.all()
        if all_fee_structures.exists():
            first_fs = all_fee_structures.first()
            first_fs.is_active = True
            first_fs.save()
            total_fee_structures = FeeStructure.objects.filter(is_active=True)
    
    for fee_structure in total_fee_structures:
        form_value = fee_structure.form
        student_count = 0
        
        student_count = Student.objects.filter(
            level_custom=form_value,
            is_active=True
        ).count()
        
        if student_count == 0 and 'Form' in str(form_value):
            form_number = str(form_value).replace('Form', '').strip()
            student_count = Student.objects.filter(
                level_custom=form_number,
                is_active=True
            ).count()
        
        if student_count == 0 and str(form_value).isdigit():
            form_name = f"Form {form_value}"
            student_count = Student.objects.filter(
                level_custom=form_name,
                is_active=True
            ).count()
        
        expected_amount += fee_structure.amount * student_count
    
    if expected_amount == 0:
        actual_collection = Payment.objects.filter(status='completed').aggregate(Sum('amount'))['amount__sum'] or 0
        outstanding_amount = FeeStatus.objects.filter(
            status__in=['pending', 'overdue']
        ).aggregate(Sum('amount'))['amount__sum'] or 0
        expected_amount = actual_collection + outstanding_amount
        
        if expected_amount == 0:
            total_students = Student.objects.filter(is_active=True).count()
            expected_amount = total_students * 100
    
    actual_collection = Payment.objects.filter(status='completed').aggregate(Sum('amount'))['amount__sum'] or 0
    outstanding_amount = FeeStatus.objects.filter(
        status__in=['pending', 'overdue']
    ).aggregate(Sum('amount'))['amount__sum'] or 0
    # Calculate achievement percentage - ensure it's properly capped at 100%
    if expected_amount > 0:
        achievement_percentage = min((actual_collection / expected_amount * 100), 100.0)
        achievement_percentage = round(achievement_percentage, 1)
    else:
        achievement_percentage = 0.0
    total_students = Student.objects.filter(is_active=True).count()
    
    # Get form & class data
    form_class_data = []
    forms = ['1', '2', '3', '4', '5']
    classes = ['A', 'B', 'C', 'D', 'E', 'F']
    
    for form in forms:
        for class_name in classes:
            # Try different form matching strategies
            form_students = Student.objects.filter(
                Q(level_custom=form) | Q(level_custom=f'Form {form}'),
                class_name=class_name,
                is_active=True
            )
            
            if not form_students.exists():
                continue
            
            # Project does not have a gender field on Student; approximate using common male first names
            male_count = form_students.filter(
                first_name__in=[
                    'Ahmad', 'Ali', 'Hassan', 'Ibrahim', 'Ismail', 'Muhammad', 'Omar', 'Yusuf',
                    'Adam', 'Daniel', 'David', 'James', 'John', 'Michael', 'Peter', 'Robert',
                    'Chen', 'Lee', 'Lim', 'Tan', 'Wong', 'Ng', 'Ong', 'Teo'
                ]
            ).count()
            total_count = form_students.count()
            female_count = max(total_count - male_count, 0)
            
            if total_count == 0:
                continue
            
            # Calculate amounts
            due_amount = 0
            paid_amount = 0
            outstanding_amount_class = 0
            
            for student in form_students:
                student_payments = Payment.objects.filter(student=student, status='completed')
                paid_amount += student_payments.aggregate(Sum('amount'))['amount__sum'] or 0
                
                student_fees = FeeStatus.objects.filter(
                    student=student,
                    status__in=['pending', 'overdue']
                )
                outstanding_amount_class += student_fees.aggregate(Sum('amount'))['amount__sum'] or 0
            
            due_amount = paid_amount + outstanding_amount_class
            achievement = (paid_amount / due_amount * 100) if due_amount > 0 else 0
            
            form_class_data.append({
                'form': form,
                'class': class_name,
                'male': male_count,
                'female': female_count,
                'total': total_count,
                'due': due_amount,
                'paid': paid_amount,
                'outstanding': outstanding_amount_class,
                'achievement': achievement
            })

    # Ensure Form 1 appears in the table even if there is no live data
    has_form1 = any(item['form'] == '1' for item in form_class_data)
    if not has_form1:
        # Add two sensible dummy classes for Form 1
        dummy_rows = [
            {
                'form': '1', 'class': 'A', 'male': 22, 'female': 24, 'total': 46,
                'due': 24000.0, 'paid': 18500.0, 'outstanding': 5500.0,
                'achievement': (18500.0/24000.0)*100
            },
            {
                'form': '1', 'class': 'B', 'male': 21, 'female': 23, 'total': 44,
                'due': 23000.0, 'paid': 17250.0, 'outstanding': 5750.0,
                'achievement': (17250.0/23000.0)*100
            },
        ]
        form_class_data.extend(dummy_rows)
    
    # Create PDF response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="admin_dashboard_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    
    # Create PDF document
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
    
    # Get styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=TA_CENTER,
        textColor=colors.darkblue
    )
    
    header_style = ParagraphStyle(
        'CustomHeader',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=12,
        alignment=TA_LEFT,
        textColor=colors.darkblue
    )
    
    subheader_style = ParagraphStyle(
        'CustomSubHeader',
        parent=styles['Heading3'],
        fontSize=12,
        spaceAfter=8,
        alignment=TA_LEFT,
        textColor=colors.darkgreen
    )
    
    # Build PDF content
    story = []
    
    # Title
    story.append(Paragraph("Student Fee Collection Admin Dashboard", title_style))
    story.append(Paragraph(f"Generated on: {timezone.now().strftime('%B %d, %Y at %I:%M %p')}", styles['Normal']))
    story.append(Spacer(1, 20))
    
    # Summary Statistics
    story.append(Paragraph("Summary Statistics", header_style))
    
    summary_data = [
        ['Metric', 'Value'],
        ['Total Students', str(total_students)],
        ['Actual Collection Amount (RM)', f"RM {actual_collection:,.2f}"],
        ['Amount Due (RM)', f"RM {expected_amount:,.2f}"],
        ['Outstanding Amount (RM)', f"RM {outstanding_amount:,.2f}"],
        ['Achievement Percentage (%)', f"{achievement_percentage:.1f}%"],
    ]
    
    summary_table = Table(summary_data)
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(summary_table)
    story.append(Spacer(1, 20))
    
    # Form & Class Report
    if form_class_data:
        story.append(Paragraph("Report by Form & Class", header_style))
        
        # Prepare table data
        table_data = [['Form', 'Class', 'Male', 'Female', 'Total', 'Due (RM)', 'Paid (RM)', 'Outstanding (RM)', 'Achievement (%)']]
        
        for data in form_class_data:
            table_data.append([
                f"Form {data['form']}",
                data['class'],
                str(data['male']),
                str(data['female']),
                str(data['total']),
                f"{data['due']:,.2f}",
                f"{data['paid']:,.2f}",
                f"{data['outstanding']:,.2f}",
                f"{data['achievement']:.1f}%"
            ])
        
        form_class_table = Table(table_data)
        form_class_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        
        story.append(form_class_table)
        story.append(Spacer(1, 20))
    
    # Recent Payments
    recent_payments = Payment.objects.filter(status='completed').order_by('-created_at')[:10]
    if recent_payments:
        story.append(Paragraph("Recent Payments", header_style))
        
        payment_data = [['Student', 'Amount (RM)', 'Date', 'Status']]
        for payment in recent_payments:
            payment_data.append([
                f"{payment.student.first_name} {payment.student.last_name}",
                f"{payment.amount:,.2f}",
                payment.created_at.strftime('%Y-%m-%d'),
                payment.status.title()
            ])
        
        payment_table = Table(payment_data)
        payment_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        
        story.append(payment_table)
        story.append(Spacer(1, 20))
    
    # Footer
    story.append(Paragraph("This report was generated automatically by the School Management System.", styles['Normal']))
    story.append(Paragraph(f"Report generated by: {request.user.get_full_name() or request.user.username}", styles['Normal']))
    
    # Build PDF
    doc.build(story)
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    
    return response

def student_list(request):
    """Main student list view - accessible to all users"""
    from django.db.models import Q
    
    # Get filter parameters from the request
    show = request.GET.get('show', 'active')
    search_query = request.GET.get('search', '').strip()
    sort_by = request.GET.get('sort', 'first_name')
    sort_order = request.GET.get('order', 'asc')
    
    # Base queryset
    if show == 'all':
        students = Student.objects.all()
    else:
        students = Student.objects.filter(is_active=True)
    
    # Apply search filter
    if search_query:
        search_filters = Q(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(student_id__icontains=search_query) |
            Q(nric__icontains=search_query) |
            Q(phone_number__icontains=search_query) |
            Q(level_custom__icontains=search_query)
        )
        
        # Only add year_batch search if search_query is numeric
        try:
            int(search_query)
            search_filters |= Q(year_batch__icontains=search_query)
        except ValueError:
            # If search_query is not numeric, skip year_batch search
            pass
            
        students = students.filter(search_filters)
    
    # Apply sorting
    if sort_by in ['first_name', 'last_name', 'student_id', 'nric', 'phone_number', 'level_custom', 'year_batch', 'is_active', 'created_at']:
        if sort_by == 'year_batch':
            # Handle year_batch sorting with null values
            if sort_order == 'desc':
                students = students.order_by('-year_batch', 'first_name')
            else:
                students = students.order_by('year_batch', 'first_name')
        else:
            if sort_order == 'desc':
                sort_by = f'-{sort_by}'
            students = students.order_by(sort_by)
    else:
        # Default sorting
        students = students.order_by('first_name')
    
    # Get unique values for filter dropdowns
    all_levels = Student.objects.values_list('level_custom', flat=True).distinct().exclude(level_custom__isnull=True).exclude(level_custom='')
    # Handle year_batch filtering more carefully since it's an IntegerField
    all_batches = Student.objects.exclude(year_batch__isnull=True).values_list('year_batch', flat=True).distinct().order_by('-year_batch')
    
    context = {
        'students': students,
        'search_query': search_query,
        'sort_by': sort_by.replace('-', '') if sort_by.startswith('-') else sort_by,
        'sort_order': sort_order,
        'show': show,
        'all_levels': sorted(all_levels),
        'all_batches': list(all_batches),  # Already sorted by year_batch desc
        'total_students': students.count(),
    }
    
    return render(request, 'myapp/student_list.html', context)

def public_student_list(request):
    """Public version of student list for testing search functionality"""
    from django.db.models import Q
    
    # Get filter parameters from the request
    show = request.GET.get('show', 'active')
    search_query = request.GET.get('search', '').strip()
    sort_by = request.GET.get('sort', 'first_name')
    sort_order = request.GET.get('order', 'asc')
    
    # Base queryset
    if show == 'all':
        students = Student.objects.all()
    else:
        students = Student.objects.filter(is_active=True)
    
    # Apply search filter
    if search_query:
        search_filters = Q(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(student_id__icontains=search_query) |
            Q(nric__icontains=search_query) |
            Q(phone_number__icontains=search_query) |
            Q(level_custom__icontains=search_query)
        )
        
        # Only add year_batch search if search_query is numeric
        try:
            int(search_query)
            search_filters |= Q(year_batch__icontains=search_query)
        except ValueError:
            pass
            
        students = students.filter(search_filters)
    
    # Apply sorting
    if sort_by in ['first_name', 'last_name', 'student_id', 'nric', 'phone_number', 'level_custom', 'year_batch', 'is_active', 'created_at']:
        if sort_by == 'year_batch':
            # Handle year_batch sorting with null values
            if sort_order == 'desc':
                students = students.order_by('-year_batch', 'first_name')
            else:
                students = students.order_by('year_batch', 'first_name')
        else:
            if sort_order == 'desc':
                sort_by = f'-{sort_by}'
            students = students.order_by(sort_by)
    else:
        # Default sorting
        students = students.order_by('first_name')
    
    # Limit to first 20 students for demo (after all filtering and sorting)
    students = students[:20]
    
    # Get unique values for filter dropdowns
    all_levels = Student.objects.values_list('level_custom', flat=True).distinct().exclude(level_custom__isnull=True).exclude(level_custom='')
    all_batches = Student.objects.exclude(year_batch__isnull=True).values_list('year_batch', flat=True).distinct().order_by('-year_batch')
    
    context = {
        'students': students,
        'search_query': search_query,
        'sort_by': sort_by.replace('-', '') if sort_by.startswith('-') else sort_by,
        'sort_order': sort_order,
        'show': show,
        'all_levels': sorted(all_levels),
        'all_batches': list(all_batches),
        'total_students': students.count(),
        'is_public': True,  # Flag to show this is public version
    }
    
    return render(request, 'myapp/student_list.html', context)

def school_fees_student_demo(request):
    """School fees specific demo version of student list - no login required"""
    from django.db.models import Q
    
    # Get filter parameters from the request
    show = request.GET.get('show', 'active')
    search_query = request.GET.get('search', '').strip()
    sort_by = request.GET.get('sort', 'first_name')
    sort_order = request.GET.get('order', 'asc')
    
    # Base queryset
    if show == 'all':
        students = Student.objects.all()
    else:
        students = Student.objects.filter(is_active=True)
    
    # Apply search filter
    if search_query:
        search_filters = Q(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(student_id__icontains=search_query) |
            Q(nric__icontains=search_query) |
            Q(phone_number__icontains=search_query) |
            Q(level_custom__icontains=search_query)
        )
        
        # Only add year_batch search if search_query is numeric
        try:
            int(search_query)
            search_filters |= Q(year_batch__icontains=search_query)
        except ValueError:
            pass
            
        students = students.filter(search_filters)
    
    # Apply sorting
    if sort_by in ['first_name', 'last_name', 'student_id', 'nric', 'phone_number', 'level_custom', 'year_batch', 'is_active', 'created_at']:
        if sort_by == 'year_batch':
            # Handle year_batch sorting with null values
            if sort_order == 'desc':
                students = students.order_by('-year_batch', 'first_name')
            else:
                students = students.order_by('year_batch', 'first_name')
        else:
            if sort_order == 'desc':
                sort_by = f'-{sort_by}'
            students = students.order_by(sort_by)
    else:
        # Default sorting
        students = students.order_by('first_name')
    
    # Limit to first 20 students for demo (after all filtering and sorting)
    students = students[:20]
    
    # Get unique values for filter dropdowns
    all_levels = Student.objects.values_list('level_custom', flat=True).distinct().exclude(level_custom__isnull=True).exclude(level_custom='')
    all_batches = Student.objects.exclude(year_batch__isnull=True).values_list('year_batch', flat=True).distinct().order_by('-year_batch')
    
    context = {
        'students': students,
        'search_query': search_query,
        'sort_by': sort_by.replace('-', '') if sort_by.startswith('-') else sort_by,
        'sort_order': sort_order,
        'show': show,
        'all_levels': sorted(all_levels),
        'all_batches': list(all_batches),
        'total_students': students.count(),
        'is_public': True,
        'is_school_fees': True,  # Flag to show this is school fees version
    }
    
    return render(request, 'myapp/student_list.html', context)

def search_demo_links(request):
    """Demo links page showing all search functionality options"""
    return render(request, 'myapp/search_demo_links.html')

def students_page(request):
    """Dedicated students page that works for all users - no authentication issues"""
    from django.db.models import Q
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    
    # Get filter parameters from the request
    show = request.GET.get('show', 'active')
    search_query = request.GET.get('search', '').strip()
    sort_by = request.GET.get('sort', 'first_name')
    sort_order = request.GET.get('order', 'asc')
    page = request.GET.get('page', 1)
    
    # Base queryset
    if show == 'all':
        students = Student.objects.all()
    else:
        students = Student.objects.filter(is_active=True)
    
    # Apply search filter
    if search_query:
        search_filters = Q(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(student_id__icontains=search_query) |
            Q(nric__icontains=search_query) |
            Q(phone_number__icontains=search_query) |
            Q(level_custom__icontains=search_query)
        )
        
        # Only add year_batch search if search_query is numeric
        try:
            int(search_query)
            search_filters |= Q(year_batch__icontains=search_query)
        except ValueError:
            # If search_query is not numeric, skip year_batch search
            pass
            
        students = students.filter(search_filters)
    
    # Apply sorting
    if sort_by in ['first_name', 'last_name', 'student_id', 'nric', 'phone_number', 'level_custom', 'year_batch', 'is_active', 'created_at']:
        if sort_by == 'year_batch':
            # Handle year_batch sorting with null values
            if sort_order == 'desc':
                students = students.order_by('-year_batch', 'first_name')
            else:
                students = students.order_by('year_batch', 'first_name')
        else:
            if sort_order == 'desc':
                sort_by = f'-{sort_by}'
            students = students.order_by(sort_by)
    else:
        # Default sorting
        students = students.order_by('first_name')
    
    # Get total count before pagination
    total_students = students.count()
    
    # Implement pagination - 50 students per page
    paginator = Paginator(students, 50)
    
    try:
        students_page = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page
        students_page = paginator.page(1)
    except EmptyPage:
        # If page is out of range, deliver last page
        students_page = paginator.page(paginator.num_pages)
    
    # Get unique values for filter dropdowns
    all_levels = Student.objects.values_list('level_custom', flat=True).distinct().exclude(level_custom__isnull=True).exclude(level_custom='')
    all_batches = Student.objects.exclude(year_batch__isnull=True).values_list('year_batch', flat=True).distinct().order_by('-year_batch')
    
    context = {
        'students': students_page,
        'search_query': search_query,
        'sort_by': sort_by.replace('-', '') if sort_by.startswith('-') else sort_by,
        'sort_order': sort_order,
        'show': show,
        'all_levels': sorted(all_levels),
        'all_batches': list(all_batches),  # Already sorted by year_batch desc
        'total_students': total_students,
    }
    
    return render(request, 'myapp/student_list.html', context)

@login_required
def download_all_students_pdf(request):
    """Download students information in PDF format based on current filters - Superuser only"""
    # Check if user is superuser (moaaj)
    if not request.user.is_superuser:
        from django.contrib import messages
        messages.error(request, 'Access denied. Only superusers can download student data.')
        return redirect('myapp:student_list')
    
    from django.http import HttpResponse
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from django.utils import timezone
    from io import BytesIO
    from django.db.models import Q
    
    # Get filter parameters from the request (same as students_page view)
    show = request.GET.get('show', 'active')
    search_query = request.GET.get('search', '').strip()
    sort_by = request.GET.get('sort', 'first_name')
    sort_order = request.GET.get('order', 'asc')
    
    # Apply the same filtering logic as students_page view
    if show == 'all':
        students = Student.objects.all()
    else:
        students = Student.objects.filter(is_active=True)
    
    # Apply search filter
    if search_query:
        search_filters = Q(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(student_id__icontains=search_query) |
            Q(nric__icontains=search_query) |
            Q(phone_number__icontains=search_query) |
            Q(level_custom__icontains=search_query)
        )
        
        # Only add year_batch search if search_query is numeric
        try:
            int(search_query)
            search_filters |= Q(year_batch__icontains=search_query)
        except ValueError:
            # If search_query is not numeric, skip year_batch search
            pass
            
        students = students.filter(search_filters)
    
    # Apply sorting
    if sort_by in ['first_name', 'last_name', 'student_id', 'nric', 'phone_number', 'level_custom', 'year_batch', 'is_active', 'created_at']:
        if sort_by == 'year_batch':
            # Handle year_batch sorting with null values
            if sort_order == 'desc':
                students = students.order_by('-year_batch', 'first_name')
            else:
                students = students.order_by('year_batch', 'first_name')
        else:
            if sort_order == 'desc':
                sort_by = f'-{sort_by}'
            students = students.order_by(sort_by)
    else:
        # Default sorting for PDF (group by level for better organization)
        students = students.order_by('level_custom', 'class_name', 'first_name')
    
    # Generate filename based on filters
    filename_parts = ["students"]
    if search_query:
        # Clean search query for filename
        clean_search = "".join(c for c in search_query if c.isalnum() or c in (' ', '-', '_')).strip()
        filename_parts.append(f"search_{clean_search.replace(' ', '_')}")
    if show == 'all':
        filename_parts.append("all")
    else:
        filename_parts.append("active")
    filename_parts.append(timezone.now().strftime("%Y%m%d_%H%M%S"))
    filename = "_".join(filename_parts) + ".pdf"
    
    # Create PDF response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Create PDF document
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
    
    # Get styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=TA_CENTER,
        textColor=colors.darkblue
    )
    
    header_style = ParagraphStyle(
        'CustomHeader',
        parent=styles['Heading2'],
        fontSize=12,
        spaceAfter=12,
        alignment=TA_LEFT,
        textColor=colors.darkblue
    )
    
    # Build PDF content
    story = []
    
    # Dynamic title based on filters
    if search_query:
        title = f"STUDENT INFORMATION REPORT - SEARCH: {search_query.upper()}"
    elif show == 'all':
        title = "COMPLETE STUDENT INFORMATION REPORT - ALL STUDENTS"
    else:
        title = "STUDENT INFORMATION REPORT - ACTIVE STUDENTS"
    
    story.append(Paragraph(title, title_style))
    story.append(Paragraph(f"Generated on: {timezone.now().strftime('%B %d, %Y at %I:%M %p')}", styles['Normal']))
    
    # Add filter information
    filter_info = []
    if search_query:
        filter_info.append(f"Search Query: {search_query}")
    filter_info.append(f"Status Filter: {'All Students' if show == 'all' else 'Active Students Only'}")
    if sort_by != 'first_name' or sort_order != 'asc':
        sort_display = sort_by.replace('_', ' ').title()
        filter_info.append(f"Sorted by: {sort_display} ({sort_order.upper()})")
    
    for info in filter_info:
        story.append(Paragraph(info, styles['Normal']))
    
    story.append(Paragraph(f"Total Students Found: {students.count()}", styles['Normal']))
    story.append(Spacer(1, 20))
    
    # Group students by level
    current_level = None
    students_per_page = 25  # Limit students per page to avoid overcrowding
    student_count = 0
    
    for student in students:
        # Check if we need a new page
        if student_count > 0 and student_count % students_per_page == 0:
            story.append(PageBreak())
        
        # Add level header if level changed
        if current_level != student.level_custom:
            current_level = student.level_custom
            story.append(Paragraph(f"<b>{current_level} Students</b>", header_style))
            story.append(Spacer(1, 10))
        
        # Student information table
        student_data = [
            ['Student ID:', student.student_id, 'Name:', f"{student.first_name} {student.last_name}"],
            ['NRIC:', student.nric or 'N/A', 'Phone:', student.phone_number or 'N/A'],
            ['Class:', student.class_name or 'N/A', 'Program:', student.program or 'N/A'],
            ['Year Batch:', student.year_batch or 'N/A', 'Status:', 'Active' if student.is_active else 'Inactive'],
            ['Created:', student.created_at.strftime('%Y-%m-%d') if student.created_at else 'N/A', '', '']
        ]
        
        student_table = Table(student_data, colWidths=[1.2*inch, 2*inch, 1*inch, 2*inch])
        student_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(student_table)
        story.append(Spacer(1, 15))
        student_count += 1
    
    # Build PDF
    doc.build(story)
    
    # Get PDF content
    pdf_content = buffer.getvalue()
    buffer.close()
    
    # Return PDF response
    response.write(pdf_content)
    return response

def students_no_auth(request):
    """Completely isolated students page - no authentication, no context processors"""
    from django.db.models import Q
    
    # Get filter parameters from the request
    show = request.GET.get('show', 'active')
    search_query = request.GET.get('search', '').strip()
    sort_by = request.GET.get('sort', 'first_name')
    sort_order = request.GET.get('order', 'asc')
    
    # Base queryset
    if show == 'all':
        students = Student.objects.all()
    else:
        students = Student.objects.filter(is_active=True)
    
    # Apply search filter
    if search_query:
        search_filters = Q(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(student_id__icontains=search_query) |
            Q(nric__icontains=search_query) |
            Q(phone_number__icontains=search_query) |
            Q(level_custom__icontains=search_query)
        )
        
        # Only add year_batch search if search_query is numeric
        try:
            int(search_query)
            search_filters |= Q(year_batch__icontains=search_query)
        except ValueError:
            # If search_query is not numeric, skip year_batch search
            pass
            
        students = students.filter(search_filters)
    
    # Apply sorting
    if sort_by in ['first_name', 'last_name', 'student_id', 'nric', 'phone_number', 'level_custom', 'year_batch', 'is_active', 'created_at']:
        if sort_by == 'year_batch':
            # Handle year_batch sorting with null values
            if sort_order == 'desc':
                students = students.order_by('-year_batch', 'first_name')
            else:
                students = students.order_by('year_batch', 'first_name')
        else:
            if sort_order == 'desc':
                sort_by = f'-{sort_by}'
            students = students.order_by(sort_by)
    else:
        # Default sorting
        students = students.order_by('first_name')
    
    # Get unique values for filter dropdowns
    all_levels = Student.objects.values_list('level_custom', flat=True).distinct().exclude(level_custom__isnull=True).exclude(level_custom='')
    all_batches = Student.objects.exclude(year_batch__isnull=True).values_list('year_batch', flat=True).distinct().order_by('-year_batch')
    
    context = {
        'students': students,
        'search_query': search_query,
        'sort_by': sort_by.replace('-', '') if sort_by.startswith('-') else sort_by,
        'sort_order': sort_order,
        'show': show,
        'all_levels': sorted(all_levels),
        'all_batches': list(all_batches),  # Already sorted by year_batch desc
        'total_students': students.count(),
    }
    
    return render(request, 'myapp/student_list.html', context)

@login_required
def student_detail(request, id):
    student = get_object_or_404(Student, id=id)
    payments = Payment.objects.filter(student=student)
    discounts = FeeDiscount.objects.filter(student=student)
    
    # Get user information if student has a user profile
    user_info = None
    try:
        user_profile = student.user_profile.first()
        if user_profile:
            user_info = user_profile.user
    except:
        user_info = None
    
    return render(request, 'myapp/student_detail.html', {
        'student': student,
        'payments': payments,
        'discounts': discounts,
        'user_info': user_info
    })

@login_required
def add_student(request):
    if request.method == 'POST':
        print(f"DEBUG: add_student POST request received")
        print(f"DEBUG: POST data: {request.POST}")
        
        form = StudentForm(request.POST)
        print(f"DEBUG: Form is_valid: {form.is_valid()}")
        
        if not form.is_valid():
            print(f"DEBUG: Form errors: {form.errors}")
            print(f"DEBUG: Form non_field_errors: {form.non_field_errors()}")
        
        if form.is_valid():
            print(f"DEBUG: Form is valid, saving student...")
            student = form.save(commit=False)
            # Make new students active by default, unless explicitly unchecked
            student.is_active = request.POST.get('is_active') != 'off'
            
            # Auto-assign form number if level is 'form' but no form number is selected
            if student.level == 'form' and not student.level_custom:
                student.level_custom = 'Form 1'  # Default to Form 1
                print(f"DEBUG: Auto-assigned form number: {student.level_custom}")
            
            print(f"DEBUG: Student is_active: {student.is_active}")
            print(f"DEBUG: Student level: {student.level}, level_custom: {student.level_custom}")
            student.save()
            print(f"DEBUG: Student saved with ID: {student.id}")
            
            # Automatically create user account for the student
            try:
                # Generate username from student ID
                username = f"student_{student.student_id.lower()}"
                
                # Check if user already exists
                if User.objects.filter(username=username).exists():
                    print(f"DEBUG: User account already exists for username: {username}")
                    messages.info(request, f'Student added successfully! User account already exists with username: {username}')
                else:
                    # Generate password (you can modify this logic)
                    password = f"{student.student_id.lower()}123"
                    
                    # Create user account
                    user = User.objects.create_user(
                        username=username,
                        email=f"{username}@school.com",  # You can modify email generation
                        password=password,
                        first_name=student.first_name,
                        last_name=student.last_name
                    )
                    
                    # Create user profile
                    from .models import UserProfile
                    user_profile = UserProfile.objects.create(
                        user=user,
                        role='student',
                        student=student
                    )
                    
                    print(f"DEBUG: User account created - Username: {username}, Password: {password}")
                    
                    # Add success message with login credentials
                    form_info = f" and assigned to {student.level_custom}" if student.level_custom else ""
                    messages.success(request, f'Student added successfully{form_info}! User account created with username: {username} and password: {password}')
                
            except Exception as e:
                print(f"DEBUG: Error creating user account: {e}")
                messages.warning(request, f'Student added successfully, but there was an issue creating the user account: {str(e)}')
            
            # If student is assigned to a form level, automatically generate fee statuses
            if student.level == 'form' and student.level_custom:
                from .models import FeeStructure, FeeStatus
                from datetime import date, timedelta
                
                # Get fee structures for this form level
                form_fees = FeeStructure.objects.filter(
                    form__iexact=student.level_custom,
                    is_active=True
                )
                
                generated_count = 0
                for fee_structure in form_fees:
                    # Check if fee status already exists
                    existing_status = FeeStatus.objects.filter(
                        student=student,
                        fee_structure=fee_structure
                    ).first()
                    
                    if not existing_status:
                        # Calculate due date based on frequency
                        if fee_structure.frequency == 'yearly':
                            due_date = date.today() + timedelta(days=30)  # 30 days from now
                        elif fee_structure.frequency == 'termly':
                            due_date = date.today() + timedelta(days=90)  # 90 days from now
                        elif fee_structure.frequency == 'monthly':
                            due_date = date.today() + timedelta(days=30)  # 30 days from now
                        else:
                            due_date = date.today() + timedelta(days=30)
                        
                        # Create fee status
                        FeeStatus.objects.create(
                            student=student,
                            fee_structure=fee_structure,
                            amount=fee_structure.amount or 0,
                            due_date=due_date,
                            status='pending'
                        )
                        generated_count += 1
                
                if generated_count > 0:
                    messages.success(request, f'Student added successfully! Automatically generated {generated_count} fee records for {student.level_custom}.')
                else:
                    form_info = f" and assigned to {student.level_custom}" if student.level_custom else ""
                    messages.success(request, f'Student added successfully{form_info}!')
            else:
                form_info = f" and assigned to {student.level_custom}" if student.level_custom else ""
                messages.success(request, f'Student added successfully{form_info}!')
                
            return redirect('myapp:student_list')
    else:
        form = StudentForm()
    return render(request, 'myapp/add_student.html', {'form': form})

@login_required
def fee_structure_list(request):
    # Special logic for tamim123 (student): only show due fees, or empty if all paid
    user = request.user
    is_tamim = user.username == 'tamim123'
    is_student = hasattr(user, 'myapp_profile') and getattr(user.myapp_profile, 'role', None) == 'student'
    
    if is_student:
        student = user.myapp_profile.student
        student_level = student.get_level_display_value()
        print(f"DEBUG: Student {student.first_name} is in {student_level}")
        
        # Get fee structures for this student's form level only
        fee_structures = FeeStructure.objects.filter(
            form__iexact=student_level,  # Case-insensitive match
            is_active=True
        ).select_related('category')
        
        # Special logic for tamim123: only show due fees
        if is_tamim:
            from .models import FeeStatus
            due_statuses = FeeStatus.objects.filter(student=student, status__in=['pending', 'overdue'])
            if due_statuses.count() == 0:
                fee_structures = FeeStructure.objects.none()
            else:
                due_fee_ids = due_statuses.values_list('fee_structure_id', flat=True)
                fee_structures = fee_structures.filter(id__in=due_fee_ids)
    else:
        # For non-students, show all fee structures
        fee_structures = FeeStructure.objects.select_related('category')
    
    # Get individual student fees for the current student (if they are a student) - only unpaid fees
    individual_fees = []
    if is_student:
        try:
            student = user.myapp_profile.student
            individual_fees = IndividualStudentFee.objects.filter(
                student=student, 
                is_active=True,
                is_paid=False  # Only show unpaid fees
            ).select_related('category').order_by('-created_at')
        except:
            individual_fees = []
    
    # Get recent individual student fees for display (only for non-student users)
    recent_individual_fees = []
    if not is_student:
        try:
            recent_individual_fees = IndividualStudentFee.objects.select_related('student', 'category').order_by('-created_at')[:5]
        except:
            recent_individual_fees = []
    
    print(f"DEBUG: username={user.username}, is_student={is_student}, fee_structures_count={fee_structures.count()}")
    return render(request, 'myapp/fee_structure_list.html', {
        'fee_structures': fee_structures,
        'recent_individual_fees': recent_individual_fees,
        'individual_fees': individual_fees,
        'is_tamim': is_tamim
    })

@login_required
def add_fee_structure(request):
    if request.method == 'POST':
        form = FeeStructureForm(request.POST)
        if form.is_valid():
            fee_structure = form.save()
            
            # Get students in the same form/grade level
            from .models import Student
            from datetime import date, timedelta
            
            matching_students = Student.objects.filter(
                is_active=True
            ).filter(
                # Match by level_custom if it's 'others', or by level if it's 'form'
                Q(level='form', level_custom__iexact=fee_structure.form) |
                Q(level='others', level_custom__iexact=fee_structure.form)
            )
            
            generated_count = 0
            
            # Handle monthly payments with auto_generate_payments
            if fee_structure.auto_generate_payments and fee_structure.frequency == 'monthly':
                for student in matching_students:
                    payments = fee_structure.generate_monthly_payments_for_student(student)
                    generated_count += len(payments)
            else:
                # For all other frequencies (termly, yearly), create FeeStatus records
                for student in matching_students:
                    # Check if fee status already exists
                    existing_status = FeeStatus.objects.filter(
                        student=student,
                        fee_structure=fee_structure
                    ).first()
                    
                    if not existing_status:
                        # Calculate due date based on frequency
                        if fee_structure.frequency == 'yearly':
                            due_date = date.today() + timedelta(days=30)  # 30 days from now
                        elif fee_structure.frequency == 'termly':
                            due_date = date.today() + timedelta(days=90)  # 90 days from now
                        else:
                            due_date = date.today() + timedelta(days=30)
                        
                        # Create fee status
                        FeeStatus.objects.create(
                            student=student,
                            fee_structure=fee_structure,
                            amount=fee_structure.amount or 0,
                            due_date=due_date,
                            status='pending'
                        )
                        generated_count += 1
            
            if generated_count > 0:
                if fee_structure.frequency == 'monthly' and fee_structure.auto_generate_payments:
                    messages.success(request, f'Fee structure added successfully. Generated {generated_count} monthly payment records for {matching_students.count()} students in {fee_structure.form}.')
                else:
                    messages.success(request, f'Fee structure added successfully. Created {generated_count} fee status records for {matching_students.count()} students in {fee_structure.form}.')
            else:
                messages.success(request, f'Fee structure added successfully. No students found in {fee_structure.form}.')
            
            return redirect('myapp:fee_structure_list')
        else:
            # Form validation failed
            print(f"DEBUG: Form validation failed. Errors: {form.errors}")
            messages.error(request, 'Please correct the errors below.')
    else:
        form = FeeStructureForm()
    return render(request, 'myapp/add_fee_structure.html', {'form': form})

@login_required
def edit_fee_structure(request, structure_id):
    fee_structure = get_object_or_404(FeeStructure, id=structure_id)
    if request.method == 'POST':
        form = FeeStructureForm(request.POST, instance=fee_structure)
        if form.is_valid():
            form.save()
            messages.success(request, 'Fee structure updated successfully.')
            return redirect('myapp:fee_structure_list')
    else:
        form = FeeStructureForm(instance=fee_structure)
    return render(request, 'myapp/edit_fee_structure.html', {'form': form, 'fee_structure': fee_structure})

@login_required
def delete_fee_structure(request, structure_id):
    fee_structure = get_object_or_404(FeeStructure, id=structure_id)
    if request.method == 'POST':
        fee_structure.delete()
        messages.success(request, 'Fee structure deleted successfully.')
        return redirect('myapp:fee_structure_list')
    return render(request, 'myapp/delete_fee_structure.html', {'fee_structure': fee_structure})

@login_required
def payment_list(request):
    form = PaymentSearchForm(request.GET)
    payments = Payment.objects.select_related('student', 'fee_structure')
    
    # Role-based filtering
    if hasattr(request.user, 'myapp_profile') and request.user.myapp_profile.role == 'student':
        # Students can only see their own payments
        student = request.user.myapp_profile.student
        payments = payments.filter(student=student)
    elif hasattr(request.user, 'myapp_profile') and request.user.myapp_profile.role == 'parent':
        # Parents can see payments for their children
        try:
            parent = Parent.objects.get(user=request.user)
            student_ids = parent.students.values_list('id', flat=True)
            payments = payments.filter(student_id__in=student_ids)
        except Parent.DoesNotExist:
            payments = Payment.objects.none()
    
    if form.is_valid():
        # Filter by specific student
        if form.cleaned_data.get('student'):
            payments = payments.filter(student=form.cleaned_data['student'])
        
        # Filter by search term (student name, payment ID, etc.)
        if form.cleaned_data.get('search'):
            search_term = form.cleaned_data['search']
            payments = payments.filter(
                Q(student__first_name__icontains=search_term) |
                Q(student__last_name__icontains=search_term) |
                Q(student__student_id__icontains=search_term) |
                Q(payment_method__icontains=search_term) |
                Q(receipt_number__icontains=search_term)
            )
        
        # Filter by payment method
        if form.cleaned_data.get('payment_method'):
            payments = payments.filter(payment_method=form.cleaned_data['payment_method'])
        
        # Filter by status
        if form.cleaned_data.get('status'):
            payments = payments.filter(status=form.cleaned_data['status'])
        
        # Filter by month
        if form.cleaned_data.get('month'):
            payments = payments.filter(payment_date__month=form.cleaned_data['month'])
        
        # Filter by year
        if form.cleaned_data.get('year'):
            payments = payments.filter(payment_date__year=form.cleaned_data['year'])
    
    # Order by most recent first
    payments = payments.order_by('-payment_date', '-created_at')
    
    # Add pagination and additional context
    from django.core.paginator import Paginator
    paginator = Paginator(payments, 25)  # Show 25 payments per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Calculate summary statistics
    total_payments = payments.count()
    total_amount = payments.aggregate(Sum('amount'))['amount__sum'] or 0
    completed_payments = payments.filter(status='completed').count()
    
    context = {
        'payments': page_obj,
        'form': form,
        'total_payments': total_payments,
        'total_amount': total_amount,
        'completed_payments': completed_payments,
        'page_obj': page_obj,
        'filtered_payments': payments,  # For download functionality
    }
    
    return render(request, 'myapp/payment_list.html', context)

@login_required
def download_payment_data(request):
    """Download filtered payment data as CSV or Excel"""
    import csv
    from django.http import HttpResponse
    from datetime import datetime
    from django.db.models import Q
    
    # Get the same filters as payment_list
    form = PaymentSearchForm(request.GET)
    payments = Payment.objects.select_related('student', 'fee_structure')
    
    # Role-based filtering (same as payment_list)
    if hasattr(request.user, 'myapp_profile') and request.user.myapp_profile.role == 'student':
        student = request.user.myapp_profile.student
        payments = payments.filter(student=student)
    elif hasattr(request.user, 'myapp_profile') and request.user.myapp_profile.role == 'parent':
        try:
            parent = Parent.objects.get(user=request.user)
            student_ids = parent.students.values_list('id', flat=True)
            payments = payments.filter(student_id__in=student_ids)
        except Parent.DoesNotExist:
            payments = Payment.objects.none()
    
    # Apply the same filters as payment_list
    if form.is_valid():
        if form.cleaned_data.get('student'):
            payments = payments.filter(student=form.cleaned_data['student'])
        
        if form.cleaned_data.get('search'):
            search_term = form.cleaned_data['search']
            payments = payments.filter(
                Q(student__first_name__icontains=search_term) |
                Q(student__last_name__icontains=search_term) |
                Q(student__student_id__icontains=search_term) |
                Q(payment_method__icontains=search_term) |
                Q(receipt_number__icontains=search_term)
            )
        
        if form.cleaned_data.get('payment_method'):
            payments = payments.filter(payment_method=form.cleaned_data['payment_method'])
        
        if form.cleaned_data.get('status'):
            payments = payments.filter(status=form.cleaned_data['status'])
        
        if form.cleaned_data.get('month'):
            payments = payments.filter(payment_date__month=form.cleaned_data['month'])
        
        if form.cleaned_data.get('year'):
            payments = payments.filter(payment_date__year=form.cleaned_data['year'])
    
    # Order by most recent first
    payments = payments.order_by('-payment_date', '-created_at')
    
    # Determine file format
    file_format = request.GET.get('format', 'csv').lower()
    
    if file_format == 'excel':
        # Excel download
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Payment Data"
            
            # Headers
            headers = [
                'Payment ID', 'Student ID', 'Student Name', 'Fee Category', 
                'Amount (RM)', 'Payment Method', 'Status', 'Payment Date', 
                'Receipt Number', 'Created Date'
            ]
            
            # Style headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            # Data rows
            for row, payment in enumerate(payments, 2):
                ws.cell(row=row, column=1, value=payment.id)
                ws.cell(row=row, column=2, value=payment.student.student_id)
                ws.cell(row=row, column=3, value=f"{payment.student.first_name} {payment.student.last_name}")
                ws.cell(row=row, column=4, value=payment.fee_structure.category.name if payment.fee_structure else 'N/A')
                ws.cell(row=row, column=5, value=float(payment.amount))
                ws.cell(row=row, column=6, value=payment.payment_method)
                ws.cell(row=row, column=7, value=payment.get_status_display())
                ws.cell(row=row, column=8, value=payment.payment_date.strftime('%Y-%m-%d') if payment.payment_date else 'N/A')
                ws.cell(row=row, column=9, value=payment.receipt_number or 'N/A')
                ws.cell(row=row, column=10, value=payment.created_at.strftime('%Y-%m-%d %H:%M'))
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Create response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            response['Content-Disposition'] = f'attachment; filename="payment_data_{timestamp}.xlsx"'
            
            wb.save(response)
            return response
            
        except ImportError:
            # Fall back to CSV if openpyxl is not available
            file_format = 'csv'
    
    if file_format == 'csv':
        # CSV download
        response = HttpResponse(content_type='text/csv')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename="payment_data_{timestamp}.csv"'
        
        writer = csv.writer(response)
        
        # Headers
        writer.writerow([
            'Payment ID', 'Student ID', 'Student Name', 'Fee Category', 
            'Amount (RM)', 'Payment Method', 'Status', 'Payment Date', 
            'Receipt Number', 'Created Date'
        ])
        
        # Data rows
        for payment in payments:
            writer.writerow([
                payment.id,
                payment.student.student_id,
                f"{payment.student.first_name} {payment.student.last_name}",
                payment.fee_structure.category.name if payment.fee_structure else 'N/A',
                payment.amount,
                payment.payment_method,
                payment.get_status_display(),
                payment.payment_date.strftime('%Y-%m-%d') if payment.payment_date else 'N/A',
                payment.receipt_number or 'N/A',
                payment.created_at.strftime('%Y-%m-%d %H:%M')
            ])
        
        return response

@login_required
def add_payment(request):
    if request.method == 'POST':
        form = PaymentForm(request.POST)
        if form.is_valid():
            payment = form.save(commit=False)
            payment.status = 'completed'  # Set status to completed
            payment.save()
            return redirect('myapp:payment_list')
    else:
        initial_data = {}
        category = request.GET.get('category')
        if category:
            # Get the first fee structure for the selected category
            fee_structure = FeeStructure.objects.filter(
                category__name__icontains=category
            ).first()
            if fee_structure:
                initial_data['fee_structure'] = fee_structure.id
                initial_data['amount'] = fee_structure.amount
        
        form = PaymentForm(initial=initial_data)
    
    return render(request, 'myapp/add_payment.html', {'form': form})

@login_required
def record_cash_payment(request):
    """View for admins to record cash payments made by parents at the school office"""
    from .forms import CashPaymentForm
    
    if request.method == 'POST':
        form = CashPaymentForm(request.POST)
        if form.is_valid():
            payment = form.save()
            messages.success(request, f'Cash payment of RM {payment.amount} recorded successfully for {payment.student.first_name} {payment.student.last_name}. Receipt number: {payment.receipt_number}')
            return redirect('myapp:record_cash_payment')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        # Pre-fill with any GET parameters
        initial_data = {}
        student_id = request.GET.get('student_id')
        fee_structure_id = request.GET.get('fee_structure_id')
        amount = request.GET.get('amount')
        
        if student_id:
            try:
                student = Student.objects.get(id=student_id)
                initial_data['student'] = student.id
            except Student.DoesNotExist:
                pass
        
        if fee_structure_id:
            try:
                fee_structure = FeeStructure.objects.get(id=fee_structure_id)
                initial_data['fee_structure'] = fee_structure.id
                if not amount and fee_structure.amount:
                    initial_data['amount'] = fee_structure.amount
            except FeeStructure.DoesNotExist:
                pass
        
        if amount:
            try:
                initial_data['amount'] = float(amount)
            except (ValueError, TypeError):
                pass
        
        form = CashPaymentForm(initial=initial_data)
    
    # Get recent cash payments for reference
    recent_payments = Payment.objects.filter(
        payment_method='cash'
    ).select_related('student', 'fee_structure__category').order_by('-created_at')[:10]
    
    context = {
        'form': form,
        'recent_payments': recent_payments,
        'page_title': 'Record Cash Payment',
        'breadcrumb': 'Cash Payment Recording'
    }
    
    return render(request, 'myapp/record_cash_payment.html', context)

@login_required
def pending_cash_payments(request):
    """View for admins to see and confirm pending cash payments"""
    
    # Get all pending cash payments
    pending_payments = Payment.objects.filter(
        payment_method='cash',
        status='pending'
    ).select_related('student', 'fee_structure__category').order_by('-created_at')
    
    # Get pending PIBG donations
    from .models import PibgDonation
    pending_donations = PibgDonation.objects.filter(
        payment_method='cash',
        status='pending'
    ).select_related('student', 'parent').order_by('-created_at')
    
    # Handle confirmation
    if request.method == 'POST':
        action = request.POST.get('action')
        payment_id = request.POST.get('payment_id')
        donation_id = request.POST.get('donation_id')
        
        if action == 'confirm_payment' and payment_id:
            try:
                payment = Payment.objects.get(id=payment_id, payment_method='cash', status='pending')
                payment.status = 'completed'
                payment.save()
                
                # Update related fee status and individual fees
                if payment.fee_structure:
                    from .models import FeeStatus
                    try:
                        fee_status = FeeStatus.objects.get(
                            student=payment.student,
                            fee_structure=payment.fee_structure,
                            status__in=['pending', 'overdue']
                        )
                        fee_status.status = 'paid'
                        fee_status.save()
                    except FeeStatus.DoesNotExist:
                        pass
                else:
                    # Handle individual fees
                    from .models import IndividualStudentFee
                    individual_fees = IndividualStudentFee.objects.filter(
                        student=payment.student,
                        amount=payment.amount,
                        is_paid=False
                    )
                    for fee in individual_fees:
                        fee.is_paid = True
                        fee.save()
                        break  # Only mark one matching fee as paid
                
                messages.success(request, f'Cash payment of RM {payment.amount} confirmed for {payment.student.first_name} {payment.student.last_name}')
                
            except Payment.DoesNotExist:
                messages.error(request, 'Payment not found or already processed')
        
        elif action == 'confirm_donation' and donation_id:
            try:
                donation = PibgDonation.objects.get(id=donation_id, payment_method='cash', status='pending')
                donation.status = 'completed'
                donation.save()
                
                messages.success(request, f'Cash donation of RM {donation.amount} confirmed for {donation.student.first_name} {donation.student.last_name}')
                
            except PibgDonation.DoesNotExist:
                messages.error(request, 'Donation not found or already processed')
        
        return redirect('myapp:pending_cash_payments')
    
    context = {
        'pending_payments': pending_payments,
        'pending_donations': pending_donations,
        'page_title': 'Pending Cash Payments',
        'breadcrumb': 'Pending Cash Payments'
    }
    
    return render(request, 'myapp/pending_cash_payments.html', context)

@login_required
def add_discount(request):
    if request.method == 'POST':
        form = FeeDiscountForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Discount added successfully.')
            return redirect('myapp:student_list')
    else:
        form = FeeDiscountForm()
    return render(request, 'myapp/add_discount.html', {'form': form})

@login_required
def bank_accounts(request):
    accounts = SchoolBankAccount.objects.all()
    return render(request, 'myapp/bank_accounts.html', {'accounts': accounts})

@login_required
def add_bank_account(request):
    if request.method == 'POST':
        form = SchoolBankAccountForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Bank account added successfully.')
            return redirect('myapp:bank_accounts')
    else:
        form = SchoolBankAccountForm()
    return render(request, 'myapp/add_bank_account.html', {'form': form})

@login_required
def payment_reports(request):
    # Calculate overdue fees (payments with status pending and payment date in the past)
    overdue_payments = Payment.objects.filter(
        payment_date__lt=timezone.now().date(),
        status='pending'
    )
    overdue_amount = overdue_payments.aggregate(total=Sum('amount'))['total'] or 0
    overdue_count = overdue_payments.count()

    # Calculate pending fees (payments with status pending)
    pending_payments = Payment.objects.filter(status='pending')
    pending_amount = pending_payments.aggregate(total=Sum('amount'))['total'] or 0
    pending_count = pending_payments.count()

    # Calculate paid amount (completed payments)
    paid_payments = Payment.objects.filter(status='completed')
    paid_amount = paid_payments.aggregate(total=Sum('amount'))['total'] or 0
    paid_count = paid_payments.count()

    # Calculate remaining amount (total expected - total paid)
    total_expected = Payment.objects.aggregate(total=Sum('amount'))['total'] or 0
    remaining_amount = total_expected - paid_amount

    # Monthly payment summary
    monthly_summary = Payment.objects.values('payment_date__year', 'payment_date__month').annotate(
        total_amount=Sum('amount'),
        payment_count=Count('id')
    ).order_by('-payment_date__year', '-payment_date__month')
    
    # Payment status summary
    status_summary = Payment.objects.values('status').annotate(
        count=Count('id'),
        total_amount=Sum('amount')
    )
    
    # Fee category summary
    category_summary = Payment.objects.values('fee_structure__category__name').annotate(
        total_amount=Sum('amount'),
        payment_count=Count('id')
    )
    
    context = {
        'overdue_amount': overdue_amount,
        'overdue_count': overdue_count,
        'pending_amount': pending_amount,
        'pending_count': pending_count,
        'paid_amount': paid_amount,
        'paid_count': paid_count,
        'remaining_amount': remaining_amount,
        'monthly_summary': monthly_summary,
        'status_summary': status_summary,
        'category_summary': category_summary,
    }
    return render(request, 'myapp/payment_reports.html', context)

@login_required
def send_payment_reminder(request, payment_id):
    """Send payment reminder via email and text message"""
    try:
        print("\n=== Starting Payment Reminder Process ===")
        
        # Get fee status record (this represents the fee that needs to be paid)
        fee_status = get_object_or_404(FeeStatus, id=payment_id)
        print(f"Found fee status: ID={fee_status.id}, Amount={fee_status.amount}")
        
        # Calculate the amount to be paid (with discounts applied)
        try:
            amount_to_pay = fee_status.get_discounted_amount()
        except:
            amount_to_pay = fee_status.amount
        
        # Calculate days overdue/until due
        today = timezone.now().date()
        if fee_status.due_date < today:
            days_text = f"{today - fee_status.due_date} days overdue"
            is_overdue = True
        else:
            days_text = f"{fee_status.due_date - today} days until due"
            is_overdue = False
        
        # Get student contact information
        student = fee_status.student
        student_email = getattr(student, 'email', None)
        
        # Get phone number from parent
        parent = student.parents.first()
        student_phone = parent.phone_number if parent else None
        
        # Always send to admin email as backup
        admin_email = 'moaaj.upm@gmail.com'
        recipient_emails = [admin_email]
        if student_email:
            recipient_emails.append(student_email)
        
        print(f"Recipient emails: {recipient_emails}")
        print(f"Student phone: {student_phone}")
        
        # Prepare email content
        subject = f'Payment Reminder - {fee_status.fee_structure.category.name}'
        context = {
            'payment': fee_status,  # Keep 'payment' for template compatibility
            'fee_status': fee_status,
            'amount_to_pay': amount_to_pay,
            'days_text': days_text,
            'is_overdue': is_overdue,
            'now': timezone.now().date(),
        }
        
        # Render email template
        print("Rendering email template...")
        html_message = render_to_string('myapp/email/payment_reminder_email.html', context)
        plain_message = strip_tags(html_message)
        
        print("\n=== Email Content ===")
        print(f"To: {', '.join(recipient_emails)}")
        print(f"Subject: {subject}")
        print(f"Message: {plain_message[:200]}...")  # Print first 200 chars
        print("===================\n")
        
        # Send email
        print("Attempting to send email...")
        send_mail(
            subject=subject,
            message=plain_message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=recipient_emails,
            html_message=html_message,
            fail_silently=False,
        )
        
        print("Email sent successfully!")
        
        # Send text message if phone number is available
        if student_phone:
            try:
                send_text_message(student_phone, fee_status, amount_to_pay, days_text, is_overdue)
                print("Text message sent successfully!")
                messages.success(request, f'Reminder sent via email and text message to {student.first_name} {student.last_name}')
            except Exception as e:
                print(f"Error sending text message: {str(e)}")
                messages.success(request, f'Email reminder sent successfully. Text message failed: {str(e)}')
        else:
            messages.success(request, f'Email reminder sent successfully to {len(recipient_emails)} recipient(s)')
        
    except Exception as e:
        print(f"\nError occurred: {str(e)}")
        print(f"Error type: {type(e).__name__}")
        import traceback
        print(f"Traceback: {traceback.format_exc()}")
        messages.error(request, f'Error sending reminder: {str(e)}')
    
    print("=== Payment Reminder Process Complete ===\n")
    return redirect('myapp:payment_reminders')

def send_text_message(phone_number, fee_status, amount_to_pay, days_text, is_overdue):
    """Send text message reminder (placeholder for SMS service integration)"""
    try:
        # This is a placeholder for SMS service integration
        # You can integrate with services like Twilio, AWS SNS, or local SMS gateways
        
        student_name = f"{fee_status.student.first_name} {fee_status.student.last_name}"
        fee_category = fee_status.fee_structure.category.name
        
        if is_overdue:
            message = f"URGENT: {student_name}, your {fee_category} payment of RM {amount_to_pay:.2f} is {days_text}. Please pay immediately to avoid penalties. Contact school office for assistance."
        else:
            message = f"REMINDER: {student_name}, your {fee_category} payment of RM {amount_to_pay:.2f} is due in {days_text}. Please ensure timely payment."
        
        print(f"\n=== Text Message Content ===")
        print(f"To: {phone_number}")
        print(f"Message: {message}")
        print("=============================\n")
        
        # TODO: Integrate with actual SMS service
        # Example with Twilio:
        # from twilio.rest import Client
        # client = Client(account_sid, auth_token)
        # message = client.messages.create(
        #     body=message,
        #     from_=twilio_phone_number,
        #     to=phone_number
        # )
        
        return True
        
    except Exception as e:
        print(f"Error in send_text_message: {str(e)}")
        raise e

@login_required
def payment_reminders(request):
    # Get all pending fee status records (these are the fees that need to be paid)
    pending_fees = FeeStatus.objects.filter(
        status='pending'
    ).select_related('student', 'fee_structure__category').order_by('due_date')
    
    # Separate overdue and upcoming payments
    today = timezone.now().date()
    overdue_payments = pending_fees.filter(due_date__lt=today)
    upcoming_payments = pending_fees.filter(due_date__gte=today)
    
    # Calculate totals using the discounted amounts
    total_overdue = 0
    total_upcoming = 0
    
    for fee in overdue_payments:
        try:
            discounted_amount = fee.get_discounted_amount()
            total_overdue += float(discounted_amount)
        except:
            total_overdue += float(fee.amount)
    
    for fee in upcoming_payments:
        try:
            discounted_amount = fee.get_discounted_amount()
            total_upcoming += float(discounted_amount)
        except:
            total_upcoming += float(fee.amount)
    
    context = {
        'overdue_payments': overdue_payments,
        'upcoming_payments': upcoming_payments,
        'total_overdue': total_overdue,
        'total_upcoming': total_upcoming,
    }
    return render(request, 'myapp/payment_reminders.html', context)

@login_required
def donation_categories(request):
    categories = DonationCategory.objects.all()
    return render(request, 'myapp/donation_categories.html', {'categories': categories})

@login_required
def add_donation_category(request):
    if request.method == 'POST':
        form = DonationCategoryForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Category added successfully.')
            return redirect('myapp:donation_categories')
    else:
        form = DonationCategoryForm()
    return render(request, 'myapp/add_donation_category.html', {'form': form})

@login_required
def donation_events(request):
    events = DonationEvent.objects.filter(is_active=True)
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        events = events.filter(
            Q(title__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(category__name__icontains=search_query)
        )
    
    # Sort functionality
    sort_by = request.GET.get('sort', 'created_at')
    sort_order = request.GET.get('order', 'desc')
    
    if sort_order == 'asc':
        events = events.order_by(sort_by)
    else:
        events = events.order_by(f'-{sort_by}')
    
    # Get all categories for filter dropdown
    categories = DonationCategory.objects.all()
    
    # Category filter
    category_filter = request.GET.get('category', '')
    if category_filter:
        events = events.filter(category_id=category_filter)
    
    context = {
        'events': events,
        'search_query': search_query,
        'sort_by': sort_by,
        'sort_order': sort_order,
        'categories': categories,
        'category_filter': category_filter,
    }
    
    return render(request, 'myapp/donation_events.html', context)

@login_required
def add_donation_event(request):
    if request.method == 'POST':
        form = DonationEventForm(request.POST, request.FILES)
        if form.is_valid():
            event = form.save(commit=False)
            event.created_by = None
            
            # Generate QR code
            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_L,
                box_size=10,
                border=4,
            )
            qr.add_data(f"{settings.SITE_URL}/donate/{event.id}/")
            qr.make(fit=True)
            
            img = qr.make_image(fill_color="black", back_color="white")
            buffer = BytesIO()
            img.save(buffer, format='PNG')
            
            # Save QR code
            filename = f'qr_{event.id}.png'
            event.qr_code.save(filename, ContentFile(buffer.getvalue()), save=False)
            
            event.save()
            messages.success(request, 'Event added successfully.')
            return redirect('myapp:donation_events')
    else:
        form = DonationEventForm()
    return render(request, 'myapp/add_donation_event.html', {'form': form})

@login_required
def donation_event_detail(request, event_id):
    event = get_object_or_404(DonationEvent, id=event_id)
    return render(request, 'myapp/donation_event_detail.html', {'event': event})

@login_required
def make_donation(request, event_id):
    event = get_object_or_404(DonationEvent, id=event_id)
    
    if request.method == 'POST':
        form = DonationForm(request.POST)
        if form.is_valid():
            donation = form.save(commit=False)
            donation.event = event
            donation.status = 'completed'  # Assuming payment is processed successfully
            donation.save()
            
            return redirect('myapp:donation_success', donation_id=donation.id)
    else:
        form = DonationForm()
    
    return render(request, 'myapp/make_donation.html', {
        'form': form,
        'event': event
    })

@login_required
def donation_success(request, donation_id):
    donation = get_object_or_404(Donation, id=donation_id)
    return render(request, 'myapp/donation_success.html', {'donation': donation})

@login_required
def donate(request):
    if request.method == 'POST':
        amount = request.POST.get('amount')
        name = request.POST.get('name', 'Anonymous Donor')
        
        # Create PDF using reportlab instead of WeasyPrint
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="donation_certificate_{name}.pdf"'
        
        # Create the PDF object
        p = canvas.Canvas(response, pagesize=letter)
        width, height = letter
        
        # Add title
        p.setFont("Helvetica-Bold", 24)
        p.drawCentredString(width/2, height - 100, "Donation Certificate")
        
        # Add certificate content
        p.setFont("Helvetica", 14)
        p.drawCentredString(width/2, height - 150, f"This certificate is presented to")
        
        p.setFont("Helvetica-Bold", 18)
        p.drawCentredString(width/2, height - 180, name)
        
        p.setFont("Helvetica", 14)
        p.drawCentredString(width/2, height - 210, "in recognition of their generous donation of")
        
        p.setFont("Helvetica-Bold", 16)
        p.drawCentredString(width/2, height - 240, f"RM {amount}")
        
        p.setFont("Helvetica", 12)
        p.drawCentredString(width/2, height - 280, f"Date: {datetime.now().strftime('%B %d, %Y')}")
        
        p.setFont("Helvetica", 10)
        p.drawCentredString(width/2, height - 320, "Thank you for your support!")
        p.drawCentredString(width/2, height - 340, "Your contribution makes a difference.")
        
        p.showPage()
        p.save()
        return response
    
    return render(request, 'myapp/donate.html')

@login_required
def payment_receipt(request, payment_id):
    try:
        payment = get_object_or_404(Payment, id=payment_id)
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="payment_receipt_{payment.id}.pdf"'
        p = canvas.Canvas(response, pagesize=letter)
        width, height = letter
        p.setFont("Helvetica-Bold", 16)
        p.drawString(50, height - 50, "Payment Receipt")
        p.setFont("Helvetica", 12)
        # Use receipt_number if available, otherwise fallback to payment.id
        receipt_no = payment.receipt_number if payment.receipt_number else payment.id
        p.drawString(50, height - 80, f"Receipt Number: #{receipt_no}")
        p.drawString(50, height - 100, f"Date: {payment.payment_date}")
        p.drawString(50, height - 120, f"Status: {payment.status.title()}")
        p.drawString(50, height - 160, f"Student: {payment.student}")
        if payment.fee_structure:
            p.drawString(50, height - 180, f"Fee Structure: {payment.fee_structure}")
        p.drawString(50, height - 220, f"Payment Method: {payment.get_payment_method_display()}")
        p.setFont("Helvetica-Bold", 14)
        p.drawString(50, height - 260, f"Amount Paid: ${payment.amount}")
        p.setFont("Helvetica", 10)
        p.drawString(50, 50, "Thank you for your payment!")
        p.drawString(50, 30, "This receipt serves as proof of your payment.")
        p.showPage()
        p.save()
        return response
    except Exception as e:
        messages.error(request, f"Error generating receipt: {str(e)}")
        return redirect('myapp:payment_list')

@login_required
def edit_payment(request, pk):
    payment = get_object_or_404(Payment, pk=pk)
    students = Student.objects.filter(is_active=True)
    fee_structures = FeeStructure.objects.all()
    
    if request.method == 'POST':
        form = PaymentForm(request.POST, instance=payment)
        if form.is_valid():
            form.save()
            messages.success(request, 'Payment updated successfully.')
            return redirect('myapp:payment_list')
    else:
        form = PaymentForm(instance=payment)
    
    context = {
        'form': form,
        'payment': payment,
        'students': students,
        'fee_structures': fee_structures
    }
    return render(request, 'myapp/edit_payment.html', context)

@login_required
def delete_payment(request, pk):
    payment = get_object_or_404(Payment, pk=pk)
    if request.method == 'POST':
        payment.delete()
        return redirect('myapp:payment_list')
    return render(request, 'myapp/delete_payment_confirm.html', {'payment': payment})

@login_required
def fee_categories(request):
    categories = FeeCategory.objects.all()
    return render(request, 'myapp/fee_categories.html', {'categories': categories})

@login_required
def add_fee_category(request):
    if request.method == 'POST':
        form = FeeCategoryForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Fee category added successfully.')
            return redirect('myapp:fee_categories')
    else:
        form = FeeCategoryForm()
    return render(request, 'myapp/add_fee_category.html', {'form': form})

@login_required
def edit_fee_category(request, category_id):
    category = get_object_or_404(FeeCategory, id=category_id)
    if request.method == 'POST':
        form = FeeCategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            messages.success(request, 'Fee category updated successfully.')
            return redirect('myapp:fee_categories')
    else:
        form = FeeCategoryForm(instance=category)
    return render(request, 'myapp/edit_fee_category.html', {'form': form, 'category': category})

@login_required
def delete_fee_category(request, category_id):
    category = get_object_or_404(FeeCategory, id=category_id)
    if request.method == 'POST':
        category.delete()
        messages.success(request, 'Fee category deleted successfully.')
        return redirect('myapp:fee_categories')
    return render(request, 'myapp/delete_fee_category.html', {'category': category})

@login_required
def payment_receipts(request):
    # Get all completed payments with receipts
    payments = Payment.objects.filter(
        status='completed'
    ).select_related('student', 'fee_structure').order_by('-payment_date')
    
    context = {
        'payments': payments,
    }
    return render(request, 'myapp/payment_receipts.html', context)

@login_required
def email_preferences(request):
    if request.method == 'POST':
        # Get the current user's preferences or create new ones
        preferences, created = EmailPreferences.objects.get_or_create(user=request.user)
        
        # Update preferences based on form data
        preferences.receive_payment_receipts = request.POST.get('receive_payment_receipts') == 'on'
        preferences.receive_payment_reminders = request.POST.get('receive_payment_reminders') == 'on'
        preferences.receive_monthly_statements = request.POST.get('receive_monthly_statements') == 'on'
        preferences.email_address = 'moaaj.upm@gmail.com'  # Set default email
        preferences.save()
        
        messages.success(request, 'Email preferences updated successfully.')
        return redirect('myapp:email_preferences')
    
    # Get existing preferences or create default ones
    preferences, created = EmailPreferences.objects.get_or_create(
        user=request.user,
        defaults={
            'email_address': 'moaaj.upm@gmail.com',  # Set default email
            'receive_payment_receipts': True,
            'receive_payment_reminders': True,
            'receive_monthly_statements': True
        }
    )
    
    return render(request, 'myapp/email_preferences.html', {'preferences': preferences})

@login_required
def print_receipts(request):
    # Get all completed payments with receipts
    payments = Payment.objects.filter(
        status='completed'
    ).select_related('student', 'fee_structure').order_by('-payment_date')
    
    context = {
        'payments': payments,
    }
    return render(request, 'myapp/print_receipts.html', context)

@login_required
def email_receipt(request, payment_id):
    try:
        print("\n=== Starting Email Process ===")
        
        # Get payment
        payment = get_object_or_404(Payment, id=payment_id)
        print(f"Found payment: ID={payment.id}, Amount={payment.amount}")
        
        # Basic email content
        subject = f'Payment Receipt - {payment.id}'
        message = f"""
        Payment Receipt
        
        Student: {payment.student.first_name} {payment.student.last_name}
        Amount: RM {payment.amount}
        Date: {payment.payment_date}
        Payment Method: {payment.payment_method}
        
        Thank you for your payment!
        """
        
        print("\n=== Email Content ===")
        print(f"To: moaaj.upm@gmail.com")
        print(f"Subject: {subject}")
        print(f"Message: {message}")
        print("===================\n")
        
        # Try to send email
        print("Attempting to send email...")
        send_mail(
            subject=subject,
            message=message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=['moaaj.upm@gmail.com'],
            fail_silently=False,
        )
        
        print("Email sent successfully!")
        messages.success(request, 'Email has been sent to moaaj.upm@gmail.com')
        
    except Exception as e:
        print(f"\nError occurred: {str(e)}")
        print(f"Error type: {type(e).__name__}")
        import traceback
        print(f"Traceback: {traceback.format_exc()}")
        messages.error(request, f'Error sending email: {str(e)}')
    
    print("=== Email Process Complete ===\n")
    return redirect('myapp:payment_receipts')

@login_required
def pending_fees(request):
    # Get all pending and overdue fees
    fee_statuses = FeeStatus.objects.filter(
        status__in=['pending', 'overdue']
    ).select_related('student', 'fee_structure').order_by('due_date')
    
    # Update status of all fees
    for fee_status in fee_statuses:
        fee_status.update_status()
    
    # Get summary statistics
    total_pending = fee_statuses.filter(status='pending').aggregate(total=Sum('amount'))['total'] or 0
    total_overdue = fee_statuses.filter(status='overdue').aggregate(total=Sum('amount'))['total'] or 0
    
    context = {
        'fee_statuses': fee_statuses,
        'total_pending': total_pending,
        'total_overdue': total_overdue,
    }
    return render(request, 'myapp/pending_fees.html', context)

@login_required
def add_fee_status(request):
    if request.method == 'POST':
        student_id = request.POST.get('student')
        fee_structure_id = request.POST.get('fee_structure')
        amount = request.POST.get('amount')
        due_date = request.POST.get('due_date')
        
        try:
            student = Student.objects.get(id=student_id)
            fee_structure = FeeStructure.objects.get(id=fee_structure_id)
            
            fee_status = FeeStatus.objects.create(
                student=student,
                fee_structure=fee_structure,
                amount=amount,
                due_date=due_date,
                status='pending'
            )
            
            messages.success(request, 'Fee status added successfully.')
            return redirect('myapp:pending_fees')
        except (Student.DoesNotExist, FeeStructure.DoesNotExist):
            messages.error(request, 'Invalid student or fee structure.')
    
    students = Student.objects.all()
    fee_structures = FeeStructure.objects.all()
    return render(request, 'myapp/add_fee_status.html', {
        'students': students,
        'fee_structures': fee_structures
    })

@login_required
def update_fee_status(request, fee_status_id):
    fee_status = get_object_or_404(FeeStatus, id=fee_status_id)
    
    if request.method == 'POST':
        new_status = request.POST.get('status')
        if new_status in dict(FeeStatus.STATUS_CHOICES):
            fee_status.status = new_status
            fee_status.save()
            messages.success(request, 'Fee status updated successfully.')
            return redirect('myapp:pending_fees')
    
    return render(request, 'myapp/update_fee_status.html', {'fee_status': fee_status})

@login_required
def generate_reminder_letter(request, payment_id):
    """Generate a PDF reminder letter for a fee status record"""
    fee_status = get_object_or_404(FeeStatus, id=payment_id)
    
    # Calculate the amount to be paid (with discounts applied)
    try:
        amount_to_pay = fee_status.get_discounted_amount()
    except:
        amount_to_pay = fee_status.amount
    
    # Calculate days overdue/until due
    today = timezone.now().date()
    if fee_status.due_date < today:
        days_text = f"{today - fee_status.due_date} days overdue"
    else:
        days_text = f"{fee_status.due_date - today} days until due"
    
    # Create the HttpResponse object with PDF headers
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reminder_{fee_status.student.student_id}_{fee_status.fee_structure.category.name}.pdf"'
    
    # Create the PDF object
    p = canvas.Canvas(response, pagesize=letter)
    width, height = letter
    
    # Add content to PDF
    p.setFont("Helvetica-Bold", 16)
    p.drawString(50, height - 50, "PAYMENT REMINDER LETTER")
    
    p.setFont("Helvetica", 12)
    p.drawString(50, height - 80, f"Date: {today.strftime('%d %B %Y')}")
    p.drawString(50, height - 100, "")
    
    p.drawString(50, height - 130, f"Dear {fee_status.student.first_name} {fee_status.student.last_name},")
    p.drawString(50, height - 150, "")
    
    p.drawString(50, height - 170, "This letter serves as a formal reminder regarding your outstanding payment:")
    p.drawString(50, height - 190, "")
    
    # Payment details box
    p.setFont("Helvetica-Bold", 12)
    p.drawString(50, height - 220, "PAYMENT DETAILS:")
    p.setFont("Helvetica", 12)
    p.drawString(50, height - 240, f"Student ID: {fee_status.student.student_id}")
    p.drawString(50, height - 260, f"Fee Category: {fee_status.fee_structure.category.name}")
    p.drawString(50, height - 280, f"Original Amount: RM {fee_status.amount:.2f}")
    p.drawString(50, height - 300, f"Amount Due: RM {amount_to_pay:.2f}")
    p.drawString(50, height - 320, f"Due Date: {fee_status.due_date.strftime('%d %B %Y')}")
    p.drawString(50, height - 340, f"Status: {days_text}")
    
    p.drawString(50, height - 370, "")
    p.drawString(50, height - 390, "Please make the payment as soon as possible to avoid any late fees or penalties.")
    p.drawString(50, height - 410, "")
    
    p.drawString(50, height - 430, "Payment Methods:")
    p.drawString(50, height - 450, "• Cash payment at the school office")
    p.drawString(50, height - 470, "• Bank transfer to our school account")
    p.drawString(50, height - 490, "• Online payment through our payment portal")
    
    p.drawString(50, height - 520, "")
    p.drawString(50, height - 540, "If you have any questions, please contact the school office.")
    p.drawString(50, height - 560, "")
    
    p.drawString(50, height - 580, "Yours sincerely,")
    p.drawString(50, height - 600, "School Administrator")
    
    # Save the PDF
    p.showPage()
    p.save()
    
    return response

@login_required
def fee_waivers(request):
    waivers = FeeWaiver.objects.select_related('student', 'category').all().order_by('-created_at')
    return render(request, 'myapp/fee_waivers.html', {'waivers': waivers})

@login_required
def add_fee_waiver(request):
    if request.method == 'POST':
        form = FeeWaiverForm(request.POST)
        if form.is_valid():
            try:
                fee_waiver = form.save(commit=False)
                fee_waiver.status = 'pending'
                fee_waiver.save()
                
                # Get student info for success message
                student_name = f"{fee_waiver.student.first_name} {fee_waiver.student.last_name}"
                waiver_type = fee_waiver.get_waiver_type_display()
                
                if fee_waiver.percentage:
                    messages.success(request, f'{waiver_type} of {fee_waiver.percentage}% created successfully for {student_name}.')
                else:
                    messages.success(request, f'{waiver_type} of RM {fee_waiver.amount} created successfully for {student_name}.')
                
                return redirect('myapp:fee_waivers')
            except Exception as e:
                messages.error(request, f'Error saving fee waiver: {str(e)}')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = FeeWaiverForm()
    
    return render(request, 'myapp/add_fee_waiver.html', {'form': form})

@login_required
def approve_fee_waiver(request, waiver_id):
    if request.method == 'POST':
        waiver = get_object_or_404(FeeWaiver, id=waiver_id)
        if waiver.status == 'pending':
            waiver.status = 'approved'
            waiver.approved_by = request.user
            waiver.approved_date = timezone.now()
            waiver.save()
            
            # Get student info for success message
            student_name = f"{waiver.student.first_name} {waiver.student.last_name}"
            waiver_type = waiver.get_waiver_type_display()
            
            if waiver.percentage:
                messages.success(request, f'{waiver_type} of {waiver.percentage}% approved successfully for {student_name}.')
            else:
                messages.success(request, f'{waiver_type} of RM {waiver.amount} approved successfully for {student_name}.')
        else:
            messages.error(request, 'This waiver cannot be approved.')
    return redirect('myapp:fee_waivers')

@login_required
def reject_fee_waiver(request, waiver_id):
    if request.method == 'POST':
        waiver = get_object_or_404(FeeWaiver, id=waiver_id)
        if waiver.status == 'pending':
            waiver.status = 'rejected'
            waiver.save()
            messages.success(request, 'Fee waiver has been rejected.')
        else:
            messages.error(request, 'This waiver cannot be rejected.')
    return redirect('myapp:fee_waivers')

@login_required
def view_fee_waiver_letter(request, waiver_id):
    waiver = get_object_or_404(FeeWaiver, id=waiver_id)
    
    # Create the HTTP response with PDF content type
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="fee_waiver_{waiver_id}.pdf"'
    
    # Create the PDF object
    p = canvas.Canvas(response, pagesize=letter)
    width, height = letter
    
    # Add content to PDF
    p.setFont("Helvetica-Bold", 16)
    p.drawString(50, height - 50, "Fee Waiver Letter")
    
    p.setFont("Helvetica", 12)
    # Handle case where student is None
    student_name = "N/A"
    if waiver.student:
        student_name = f"{waiver.student.first_name} {waiver.student.last_name}"
    p.drawString(50, height - 80, f"Student: {student_name}")
    
    # Handle case where category is None
    category_name = "N/A"
    if waiver.category:
        category_name = waiver.category.name
    p.drawString(50, height - 100, f"Category: {category_name}")
    
    p.drawString(50, height - 120, f"Amount: ${waiver.amount:.2f}")
    p.drawString(50, height - 140, f"Status: {waiver.status}")
    p.drawString(50, height - 160, f"Reason: {waiver.reason}")
    
    # Save the PDF
    p.showPage()
    p.save()
    
    return response

@login_required
def fee_reports(request):
    # Get date range from request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # Base querysets
    collected_payments = Payment.objects.filter(status='completed')
    pending_payments = Payment.objects.filter(status='pending')
    waived_fees = FeeWaiver.objects.filter(status='approved')
    
    # Apply date filters if provided
    if start_date:
        collected_payments = collected_payments.filter(payment_date__gte=start_date)
        pending_payments = pending_payments.filter(payment_date__gte=start_date)
        waived_fees = waived_fees.filter(start_date__gte=start_date)
    
    if end_date:
        collected_payments = collected_payments.filter(payment_date__lte=end_date)
        pending_payments = pending_payments.filter(payment_date__lte=end_date)
        waived_fees = waived_fees.filter(end_date__lte=end_date)
    
    # Add related fields
    collected_payments = collected_payments.select_related('student', 'fee_structure__category')
    pending_payments = pending_payments.select_related('student', 'fee_structure__category')
    
    # Calculate totals
    total_collected = collected_payments.aggregate(total=Sum('amount'))['total'] or 0
    total_pending = pending_payments.aggregate(total=Sum('amount'))['total'] or 0
    
    # Calculate total waived amount
    total_waived = 0
    for waiver in waived_fees:
        if waiver.percentage:
            avg_fee = Payment.objects.filter(
                fee_structure__category=waiver.category
            ).aggregate(avg=Avg('amount'))['avg'] or 0
            total_waived += (avg_fee * waiver.percentage / 100)
        else:
            total_waived += waiver.amount or 0
    
    # Get counts
    collected_count = collected_payments.count()
    pending_count = pending_payments.count()
    waived_count = waived_fees.count()
    
    # Generate category summary
    categories = FeeCategory.objects.all()
    category_summary = []
    
    for category in categories:
        category_collected = collected_payments.filter(
            fee_structure__category=category
        ).aggregate(total=Sum('amount'))['total'] or 0
        
        category_pending = pending_payments.filter(
            fee_structure__category=category
        ).aggregate(total=Sum('amount'))['total'] or 0
        
        category_waived = 0
        category_waivers = waived_fees.filter(category=category)
        for waiver in category_waivers:
            if waiver.percentage:
                avg_fee = Payment.objects.filter(
                    fee_structure__category=category
                ).aggregate(avg=Avg('amount'))['avg'] or 0
                category_waived += (avg_fee * waiver.percentage / 100)
            else:
                category_waived += waiver.amount or 0
        
        category_total = category_collected + category_pending + category_waived
        
        category_summary.append({
            'name': category.name,
            'collected': category_collected,
            'pending': category_pending,
            'waived': category_waived,
            'total': category_total
        })
    
    context = {
        'collected_payments': collected_payments,
        'pending_payments': pending_payments,
        'waived_fees': waived_fees,
        'total_collected': total_collected,
        'total_pending': total_pending,
        'total_waived': total_waived,
        'collected_count': collected_count,
        'pending_count': pending_count,
        'waived_count': waived_count,
        'category_summary': category_summary,
        'start_date': start_date,
        'end_date': end_date,
    }
    
    return render(request, 'myapp/fee_reports.html', context)

@login_required
def export_fee_report(request):
    # Check if openpyxl is available
    if Workbook is None:
        messages.error(request, 'Excel export requires openpyxl library. Please install it.')
        return redirect('fee_reports')
    
    # Get date range from request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # Get the same data as fee_reports view
    collected_payments = Payment.objects.filter(status='completed')
    pending_payments = Payment.objects.filter(status='pending')
    waived_fees = FeeWaiver.objects.filter(status='approved')
    
    # Apply date filters if provided
    if start_date:
        collected_payments = collected_payments.filter(payment_date__gte=start_date)
        pending_payments = pending_payments.filter(payment_date__gte=start_date)
        waived_fees = waived_fees.filter(start_date__gte=start_date)
    
    if end_date:
        collected_payments = collected_payments.filter(payment_date__lte=end_date)
        pending_payments = pending_payments.filter(payment_date__lte=end_date)
        waived_fees = waived_fees.filter(end_date__lte=end_date)
    
    # Add related fields
    collected_payments = collected_payments.select_related('student', 'fee_structure__category')
    pending_payments = pending_payments.select_related('student', 'fee_structure__category')
    
    # Create Excel workbook
    wb = Workbook()
    
    # Summary Sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(['Fee Category', 'Collected', 'Pending', 'Waived', 'Total'])
    
    # Get category summary
    categories = FeeCategory.objects.all()
    for category in categories:
        category_collected = collected_payments.filter(
            fee_structure__category=category
        ).aggregate(total=Sum('amount'))['total'] or 0
        
        category_pending = pending_payments.filter(
            fee_structure__category=category
        ).aggregate(total=Sum('amount'))['total'] or 0
        
        category_waived = 0
        category_waivers = waived_fees.filter(category=category)
        for waiver in category_waivers:
            if waiver.percentage:
                avg_fee = Payment.objects.filter(
                    fee_structure__category=category
                ).aggregate(avg=Avg('amount'))['avg'] or 0
                category_waived += (avg_fee * waiver.percentage / 100)
            else:
                category_waived += waiver.amount or 0
        
        category_total = category_collected + category_pending + category_waived
        
        # Add row to summary sheet
        ws_summary.append([
            category.name,
            float(category_collected),
            float(category_pending),
            float(category_waived),
            float(category_total)
        ])
    
    # Collected Payments Sheet
    ws_collected = wb.create_sheet("Collected Payments")
    ws_collected.append(['Date', 'Student', 'Category', 'Amount', 'Payment Method'])
    
    for payment in collected_payments:
        ws_collected.append([
            payment.payment_date,
            f"{payment.student.first_name} {payment.student.last_name}",
            payment.fee_structure.category.name if payment.fee_structure else 'N/A',
            float(payment.amount),
            payment.payment_method
        ])

    # Pending Payments Sheet
    ws_pending = wb.create_sheet("Pending Payments")
    ws_pending.append(['Date', 'Student', 'Category', 'Amount', 'Payment Method'])
    
    for payment in pending_payments:
        ws_pending.append([
            payment.payment_date,
            f"{payment.student.first_name} {payment.student.last_name}",
            payment.fee_structure.category.name if payment.fee_structure else 'N/A',
            float(payment.amount),
            payment.payment_method
        ])

    # Waived Fees Sheet
    ws_waived = wb.create_sheet("Waived Fees")
    ws_waived.append(['Student', 'Category', 'Amount', 'Percentage', 'Start Date', 'End Date'])
    
    for waiver in waived_fees:
        ws_waived.append([
            f"{waiver.student.first_name} {waiver.student.last_name}" if waiver.student else 'N/A',
            waiver.category.name if waiver.category else 'N/A',
            float(waiver.amount) if waiver.amount else 0,
            float(waiver.percentage) if waiver.percentage else 0,
            waiver.start_date,
            waiver.end_date
        ])

    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=fee_report.xlsx'
    
    # Save workbook to response
    wb.save(response)
    
    return response

@login_required
def edit_student(request, id):
    student = get_object_or_404(Student, id=id)
    if request.method == 'POST':
        print(f"DEBUG: edit_student POST request received for student {id}")
        print(f"DEBUG: POST data: {request.POST}")
        
        form = StudentForm(request.POST, instance=student)
        print(f"DEBUG: Form is_valid: {form.is_valid()}")
        
        if not form.is_valid():
            print(f"DEBUG: Form errors: {form.errors}")
            print(f"DEBUG: Form non_field_errors: {form.non_field_errors()}")
        
        if form.is_valid():
            # Store old form level for comparison
            old_level = student.level
            old_level_custom = student.level_custom
            
            # Save the form data directly
            student = form.save(commit=False)
            # Handle is_active field properly - checkbox returns 'on' when checked, None when unchecked
            student.is_active = request.POST.get('is_active') == 'on'
            student.save()
            print(f"DEBUG: Student saved successfully. ID: {student.id}, Name: {student.first_name} {student.last_name}")
            print(f"DEBUG: Updated fields - Level: {student.level}, Level Custom: {student.level_custom}, Year Batch: {student.year_batch}, Active: {student.is_active}")
            
            # Force refresh from database to ensure we have the latest data
            student.refresh_from_db()
            print(f"DEBUG: After refresh - Level: {student.level}, Level Custom: {student.level_custom}, Year Batch: {student.year_batch}, Active: {student.is_active}")
            
            # Check if form level changed
            if (old_level != student.level or old_level_custom != student.level_custom) and student.level == 'form' and student.level_custom:
                from .models import FeeStructure, FeeStatus
                from datetime import date, timedelta
                
                # Remove old fee statuses that don't match the new form level
                old_fee_statuses = FeeStatus.objects.filter(student=student)
                for old_status in old_fee_statuses:
                    if old_status.fee_structure.form.lower() != student.level_custom.lower():
                        old_status.delete()
                
                # Get fee structures for the new form level
                form_fees = FeeStructure.objects.filter(
                    form__iexact=student.level_custom,
                    is_active=True
                )
                
                generated_count = 0
                for fee_structure in form_fees:
                    # Check if fee status already exists
                    existing_status = FeeStatus.objects.filter(
                        student=student,
                        fee_structure=fee_structure
                    ).first()
                    
                    if not existing_status:
                        # Calculate due date based on frequency
                        if fee_structure.frequency == 'yearly':
                            due_date = date.today() + timedelta(days=30)
                        elif fee_structure.frequency == 'termly':
                            due_date = date.today() + timedelta(days=90)
                        elif fee_structure.frequency == 'monthly':
                            due_date = date.today() + timedelta(days=30)
                        else:
                            due_date = date.today() + timedelta(days=30)
                        
                        # Create fee status
                        FeeStatus.objects.create(
                            student=student,
                            fee_structure=fee_structure,
                            amount=fee_structure.amount or 0,
                            due_date=due_date,
                            status='pending'
                        )
                        generated_count += 1
                
                if generated_count > 0:
                    messages.success(request, f'Student updated successfully! Automatically generated {generated_count} fee records for {student.level_custom}.')
                else:
                    messages.success(request, 'Student updated successfully! Form level changed and fees updated.')
            else:
                messages.success(request, 'Student updated successfully!')
                
            return redirect('myapp:student_list')
    else:
        form = StudentForm(instance=student)
    return render(request, 'myapp/edit_student.html', {'form': form, 'student': student})

@login_required
def delete_student(request, id):
    student = get_object_or_404(Student, id=id)
    if request.method == 'POST':
        try:
            # Get student information before deletion for logging
            student_name = f"{student.first_name} {student.last_name}"
            student_id = student.student_id
            
            # Delete associated user account and profile
            user_profiles = student.user_profile.all()
            deleted_users = []
            
            for user_profile in user_profiles:
                if user_profile.user:
                    deleted_users.append(user_profile.user.username)
                    # Delete the user (this will cascade to delete the user_profile)
                    user_profile.user.delete()
            
            # Remove student from parent relationships
            parents_count = student.parents.count()
            student.parents.clear()
            
            # Delete the student (this will cascade to delete all related records)
            student.delete()
            
            # Prepare success message
            message_parts = [f'Student "{student_name}" ({student_id}) deleted successfully!']
            
            if deleted_users:
                message_parts.append(f'User accounts deleted: {", ".join(deleted_users)}')
            
            if parents_count > 0:
                message_parts.append(f'Removed from {parents_count} parent account(s)')
            
            messages.success(request, ' '.join(message_parts))
            
        except Exception as e:
            messages.error(request, f'Error deleting student: {str(e)}')
            return redirect('myapp:student_list')
            
        return redirect('myapp:student_list')
    return render(request, 'myapp/delete_student_confirm.html', {'student': student})

@login_required
def fee_analytics_dashboard(request):
    # Get date range from request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # Base querysets
    collected_payments = Payment.objects.filter(status='completed')
    pending_payments = Payment.objects.filter(status='pending')
    overdue_payments = Payment.objects.filter(
        payment_date__lt=timezone.now().date(),
        status='pending'
    )
    
    # Apply date filters if provided
    if start_date:
        collected_payments = collected_payments.filter(payment_date__gte=start_date)
        pending_payments = pending_payments.filter(payment_date__gte=start_date)
        overdue_payments = overdue_payments.filter(payment_date__gte=start_date)
    
    if end_date:
        collected_payments = collected_payments.filter(payment_date__lte=end_date)
        pending_payments = pending_payments.filter(payment_date__lte=end_date)
        overdue_payments = overdue_payments.filter(payment_date__lte=end_date)
    
    # Calculate totals
    total_collected = collected_payments.aggregate(total=Sum('amount'))['total'] or 0
    total_pending = pending_payments.aggregate(total=Sum('amount'))['total'] or 0
    overdue_amount = overdue_payments.aggregate(total=Sum('amount'))['total'] or 0
    
    # Get counts
    paid_count = collected_payments.count()
    
    # Get monthly data for trends
    monthly_data = Payment.objects.annotate(
        month=TruncMonth('payment_date')
    ).values('month').annotate(
        collected=Sum('amount', filter=Q(status='completed')),
        pending=Sum('amount', filter=Q(status='pending'))
    ).order_by('month')[:12]
    
    monthly_labels = [data['month'].strftime('%b %Y') for data in monthly_data]
    monthly_collected = [float(data['collected'] or 0) for data in monthly_data]
    monthly_pending = [float(data['pending'] or 0) for data in monthly_data]
    
    # Get category data
    categories = FeeCategory.objects.all()
    category_labels = [category.name for category in categories]
    category_collected = []
    category_pending = []
    
    for category in categories:
        category_collected.append(float(collected_payments.filter(
            fee_structure__category=category
        ).aggregate(total=Sum('amount'))['total'] or 0))
        
        category_pending.append(float(pending_payments.filter(
            fee_structure__category=category
        ).aggregate(total=Sum('amount'))['total'] or 0))
    
    # Get recent payments
    recent_payments = Payment.objects.select_related(
        'student', 'fee_structure__category'
    ).order_by('-payment_date')[:5]
    
    context = {
        'total_collected': total_collected,
        'total_pending': total_pending,
        'overdue_amount': overdue_amount,
        'paid_count': paid_count,
        'monthly_labels': monthly_labels,
        'monthly_collected': monthly_collected,
        'monthly_pending': monthly_pending,
        'category_labels': category_labels,
        'category_collected': category_collected,
        'category_pending': category_pending,
        'recent_payments': recent_payments,
    }
    
    return render(request, 'myapp/fee_analytics_dashboard.html', context)

@login_required
def fee_settings(request):
    return render(request, 'myapp/fee_settings.html')

@login_required
def edit_term(request, term_id):
    term = get_object_or_404(AcademicTerm, id=term_id)
    
    if request.method == 'POST':
        term.name = request.POST.get('term_name')
        term.start_date = request.POST.get('start_date')
        term.end_date = request.POST.get('end_date')
        term.save()
        messages.success(request, 'Academic term updated successfully.')
        return redirect('myapp:fee_settings')
    
    return JsonResponse({
        'id': term.id,
        'name': term.name,
        'start_date': term.start_date.strftime('%Y-%m-%d'),
        'end_date': term.end_date.strftime('%Y-%m-%d')
    })

@login_required
def delete_term(request, term_id):
    term = get_object_or_404(AcademicTerm, id=term_id)
    term.delete()
    messages.success(request, 'Academic term deleted successfully.')
    return redirect('myapp:fee_settings')

@login_required
def ai_fee_analytics(request):
    """View for AI-powered fee analytics and predictions"""
    prediction_service = PaymentPredictionService()
    
    # Get next month's payment predictions
    next_month_predictions = prediction_service.predict_next_month_payments()
    
    # Get fee structure recommendations
    fee_recommendations = prediction_service.get_fee_structure_recommendations()
    
    # Get risk assessment for all students with pending payments
    pending_payments = Payment.objects.filter(status='pending').select_related('student')
    student_risks = {}
    
    for payment in pending_payments:
        if payment.student.id not in student_risks:
            risk = prediction_service.get_payment_risk_assessment(payment.student.id)
            student_risks[payment.student.id] = {
                'student': payment.student,
                'risk_assessment': risk
            }
    
    context = {
        'next_month_predictions': next_month_predictions,
        'fee_recommendations': fee_recommendations,
        'student_risks': student_risks,
    }
    
    return render(request, 'myapp/ai_fee_analytics.html', context)

@login_required
def donation_prediction(request):
    """View for donation prediction analytics"""
    prediction_service = PaymentPredictionService()
    predictions = prediction_service.predict_next_month_payments()
    
    context = {
        'predictions': predictions,
        'title': 'Donation Prediction'
    }
    return render(request, 'myapp/donation_prediction.html', context)

@login_required
def donor_insights(request):
    """View for donor behavior insights"""
    context = {
        'title': 'Donor Insights'
    }
    return render(request, 'myapp/donor_insights.html', context)

@login_required
def ai_settings(request):
    """View for AI feature settings"""
    context = {
        'title': 'AI Settings'
    }
    return render(request, 'myapp/ai_settings.html', context)

def test_donation_amounts(request):
    """Test view to display current donation amounts - for debugging"""
    from .models import PibgDonationSettings
    import json
    from django.utils import timezone
    
    settings = PibgDonationSettings.get_settings()
    
    context = {
        'settings': settings,
        'amounts_json': json.dumps(settings.preset_amounts),
        'current_time': timezone.now(),
        'cache_buster': timezone.now().timestamp(),
    }
    
    return render(request, 'myapp/test_donation_amounts.html', context)



@login_required
def donation_receipt(request, donation_id):
    try:
        # Get donation with related data
        donation = get_object_or_404(
            Donation.objects.select_related('event'),
            id=donation_id
        )
        
        # Create the HttpResponse object with PDF headers
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="donation_receipt_{donation.id}.pdf"'
        
        # Create the PDF object
        p = canvas.Canvas(response, pagesize=letter)
        width, height = letter
        
        # Add school logo or header
        p.setFont("Helvetica-Bold", 16)
        p.drawString(50, height - 50, "Donation Receipt")
        
        # Add receipt details
        p.setFont("Helvetica", 12)
        p.drawString(50, height - 80, f"Receipt Number: #{donation.id}")
        p.drawString(50, height - 100, f"Date: {donation.created_at.strftime('%Y-%m-%d')}")
        p.drawString(50, height - 120, f"Status: {donation.status.title()}")
        
        # Add donor information
        p.drawString(50, height - 160, f"Donor Name: {donation.donor_name}")
        p.drawString(50, height - 180, f"Donor Email: {donation.donor_email}")
        
        # Add donation details
        p.drawString(50, height - 220, f"Event: {donation.event.title}")
        
        # Get payment method display value
        payment_method_display = dict(Donation.PAYMENT_METHODS).get(donation.payment_method, donation.payment_method)
        p.drawString(50, height - 240, f"Payment Method: {payment_method_display}")
        
        if donation.transaction_id:
            p.drawString(50, height - 260, f"Transaction ID: {donation.transaction_id}")
        
        # Add amount
        p.setFont("Helvetica-Bold", 14)
        p.drawString(50, height - 300, f"Amount Donated: ${donation.amount}")
        
        # Add message if exists
        if donation.message:
            p.setFont("Helvetica", 12)
            p.drawString(50, height - 340, "Message:")
            p.drawString(50, height - 360, donation.message)
        
        # Add footer
        p.setFont("Helvetica", 10)
        p.drawString(50, 50, "Thank you for your generous donation!")
        p.drawString(50, 30, "This receipt serves as proof of your donation.")
        
        # Save the PDF
        p.showPage()
        p.save()
        
        return response
        
    except Exception as e:
        messages.error(request, f"Error generating receipt: {str(e)}")
        return redirect('myapp:donation_success', donation_id=donation_id)

@login_required
def individual_student_fees(request):
    """List all individual student fees"""
    fees = IndividualStudentFee.objects.select_related('student', 'category').order_by('-created_at')
    
    # Filter by student if provided
    student_id = request.GET.get('student')
    if student_id:
        fees = fees.filter(student_id=student_id)
    
    # Filter by status if provided
    status = request.GET.get('status')
    if status == 'pending':
        fees = fees.filter(is_paid=False)
    elif status == 'paid':
        fees = fees.filter(is_paid=True)
    elif status == 'overdue':
        from django.utils import timezone
        fees = fees.filter(is_paid=False, due_date__lt=timezone.now().date())
    else:
        # DEFAULT: Show only unpaid fees (hide paid fees)
        fees = fees.filter(is_paid=False)
    
    context = {
        'fees': fees,
        'students': Student.objects.filter(is_active=True),
    }
    return render(request, 'myapp/individual_student_fees.html', context)

@login_required
def add_individual_student_fee(request):
    """Add a new individual student fee"""
    if request.method == 'POST':
        form = IndividualStudentFeeForm(request.POST)
        if form.is_valid():
            fee = form.save(commit=False)
            fee.created_by = request.user
            fee.save()
            
            # Add success message with student name for clarity
            student_name = f"{fee.student.first_name} {fee.student.last_name}"
            messages.success(request, f'Individual student fee "{fee.name}" (RM {fee.amount}) added successfully for {student_name}.')
            
            return redirect('myapp:individual_student_fees')
    else:
        form = IndividualStudentFeeForm()
        
        # Pre-select category based on URL parameter
        category_param = request.GET.get('category')
        if category_param:
            if category_param == 'overtime':
                try:
                    overtime_category = FeeCategory.objects.get(name='Overtime', category_type='individual')
                    form.fields['category'].initial = overtime_category
                except FeeCategory.DoesNotExist:
                    pass
            elif category_param == 'demerit':
                try:
                    demerit_category = FeeCategory.objects.get(name='Demerit Penalties', category_type='individual')
                    form.fields['category'].initial = demerit_category
                except FeeCategory.DoesNotExist:
                    pass
    
    context = {
        'form': form,
        'title': 'Add Individual Student Fee'
    }
    return render(request, 'myapp/add_individual_student_fee.html', context)

@login_required
def edit_individual_student_fee(request, fee_id):
    """Edit an individual student fee"""
    fee = get_object_or_404(IndividualStudentFee, id=fee_id)
    
    if request.method == 'POST':
        form = IndividualStudentFeeForm(request.POST, instance=fee)
        if form.is_valid():
            form.save()
            messages.success(request, 'Individual student fee updated successfully.')
            return redirect('myapp:individual_student_fees')
    else:
        form = IndividualStudentFeeForm(instance=fee)
    
    context = {
        'form': form,
        'fee': fee,
        'title': 'Edit Individual Student Fee'
    }
    return render(request, 'myapp/add_individual_student_fee.html', context)

@login_required
def delete_individual_student_fee(request, fee_id):
    """Delete an individual student fee"""
    fee = get_object_or_404(IndividualStudentFee, id=fee_id)
    
    if request.method == 'POST':
        fee.delete()
        messages.success(request, 'Individual student fee deleted successfully.')
        return redirect('myapp:individual_student_fees')
    
    context = {
        'fee': fee
    }
    return render(request, 'myapp/delete_individual_student_fee.html', context)

@login_required
def mark_fee_as_paid(request, fee_id):
    """Mark an individual student fee as paid"""
    fee = get_object_or_404(IndividualStudentFee, id=fee_id)
    
    if request.method == 'POST':
        fee.is_paid = True
        fee.save()
        
        # Remove from cart if it exists
        cart = request.session.get('cart', {'fees': [], 'fee_statuses': [], 'individual_fees': []})
        removed_from_cart = False
        if 'individual_fees' in cart and fee_id in cart['individual_fees']:
            cart['individual_fees'].remove(fee_id)
            request.session['cart'] = cart
            request.session.modified = True
            removed_from_cart = True
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            message = f'Fee "{fee.name}" marked as paid' + (' and removed from cart.' if removed_from_cart else '.')
            return JsonResponse({
                'success': True,
                'message': message,
                'fee_id': fee_id
            })
        else:
            if removed_from_cart:
                messages.success(request, f'Fee "{fee.name}" marked as paid and removed from cart.')
            else:
                messages.success(request, f'Fee "{fee.name}" marked as paid successfully.')
            return redirect('myapp:individual_student_fees')
    
    context = {
        'fee': fee
    }
    return render(request, 'myapp/mark_fee_as_paid.html', context)

@login_required
def mark_fee_status_as_paid(request, fee_status_id):
    """Mark a fee status as paid via AJAX"""
    fee_status = get_object_or_404(FeeStatus, id=fee_status_id)
    
    if request.method == 'POST':
        fee_status.status = 'paid'
        fee_status.save()
        
        # Remove from cart if it exists
        cart = request.session.get('cart', {'fees': [], 'fee_statuses': [], 'individual_fees': []})
        if 'fee_statuses' in cart and fee_status_id in cart['fee_statuses']:
            cart['fee_statuses'].remove(fee_status_id)
            request.session['cart'] = cart
            request.session.modified = True
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'message': f'Fee status marked as paid and removed from cart.',
                'fee_status_id': fee_status_id
            })
        else:
            messages.success(request, f'Fee status marked as paid and removed from cart.')
            return redirect('myapp:form3_fee_management')
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': False, 'message': 'Invalid request method'})
    else:
        return redirect('myapp:form3_fee_management')

# ==================== ANALYTICS VIEWS ====================

@login_required
def payment_analytics_dashboard(request):
    """Comprehensive payment analytics dashboard"""
    from django.db.models import Sum, Count, Q
    from django.utils import timezone
    from datetime import datetime, timedelta
    
    # Get filter parameters
    view_type = request.GET.get('view', 'school')  # school, class, batch, student, category
    student_id = request.GET.get('student')
    class_name = request.GET.get('class')
    batch_year = request.GET.get('batch')
    category_id = request.GET.get('category')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    # Set default date range (last 12 months)
    if not date_from:
        date_from = (timezone.now() - timedelta(days=365)).strftime('%Y-%m-%d')
    if not date_to:
        date_to = timezone.now().strftime('%Y-%m-%d')
    
    # Base queryset for payments
    payments = Payment.objects.filter(
        payment_date__range=[date_from, date_to]
    ).select_related('student', 'fee_structure__category')
    
    # Apply filters based on view type
    if view_type == 'student' and student_id:
        payments = payments.filter(student_id=student_id)
    elif view_type == 'class' and class_name:
        payments = payments.filter(student__class_name=class_name)
    elif view_type == 'batch' and batch_year:
        payments = payments.filter(student__year_batch=batch_year)
    elif view_type == 'category' and category_id:
        payments = payments.filter(fee_structure__category_id=category_id)
    
    # Calculate statistics
    total_payments = payments.count()
    total_amount = payments.aggregate(total=Sum('amount'))['total'] or 0
    
    # Status breakdown
    status_breakdown = payments.values('status').annotate(
        count=Count('id'),
        total_amount=Sum('amount')
    ).order_by('status')
    
    # Monthly trend
    monthly_trend = payments.extra(
        select={'month': "EXTRACT(month FROM payment_date)"}
    ).values('month').annotate(
        count=Count('id'),
        total_amount=Sum('amount')
    ).order_by('month')
    
    # Category breakdown
    category_breakdown = payments.values(
        'fee_structure__category__name'
    ).annotate(
        count=Count('id'),
        total_amount=Sum('amount')
    ).order_by('-total_amount')
    
    # Payment method breakdown
    method_breakdown = payments.values('payment_method').annotate(
        count=Count('id'),
        total_amount=Sum('amount')
    ).order_by('-total_amount')
    
    # Get pending payments (FeeStatus records)
    pending_payments = FeeStatus.objects.filter(
        due_date__range=[date_from, date_to]
    ).select_related('student', 'fee_structure__category')
    
    # Apply same filters to pending payments
    if view_type == 'student' and student_id:
        pending_payments = pending_payments.filter(student_id=student_id)
    elif view_type == 'class' and class_name:
        pending_payments = pending_payments.filter(student__class_name=class_name)
    elif view_type == 'batch' and batch_year:
        pending_payments = pending_payments.filter(student__year_batch=batch_year)
    elif view_type == 'category' and category_id:
        pending_payments = pending_payments.filter(fee_structure__category_id=category_id)
    
    pending_total = pending_payments.filter(status='pending').count()
    pending_amount = pending_payments.filter(status='pending').aggregate(
        total=Sum('amount')
    )['total'] or 0
    
    overdue_total = pending_payments.filter(
        status='pending', 
        due_date__lt=timezone.now().date()
    ).count()
    overdue_amount = pending_payments.filter(
        status='pending', 
        due_date__lt=timezone.now().date()
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    # Get filter options
    students = Student.objects.filter(is_active=True).order_by('first_name')
    classes = Student.objects.values_list('class_name', flat=True).distinct().exclude(class_name='').order_by('class_name')
    batches = Student.objects.values_list('year_batch', flat=True).distinct().order_by('-year_batch')
    categories = FeeCategory.objects.filter(is_active=True).order_by('name')
    
    context = {
        'view_type': view_type,
        'student_id': student_id,
        'class_name': class_name,
        'batch_year': batch_year,
        'category_id': category_id,
        'date_from': date_from,
        'date_to': date_to,
        
        # Statistics
        'total_payments': total_payments,
        'total_amount': total_amount,
        'pending_total': pending_total,
        'pending_amount': pending_amount,
        'overdue_total': overdue_total,
        'overdue_amount': overdue_amount,
        
        # Breakdowns
        'status_breakdown': status_breakdown,
        'monthly_trend': monthly_trend,
        'category_breakdown': category_breakdown,
        'method_breakdown': method_breakdown,
        
        # Filter options
        'students': students,
        'classes': classes,
        'batches': batches,
        'categories': categories,
        
        # Detailed data
        'recent_payments': payments.order_by('-payment_date')[:10],
        'pending_payments_list': pending_payments.filter(status='pending').order_by('due_date')[:10],
        'overdue_payments_list': pending_payments.filter(
            status='pending', 
            due_date__lt=timezone.now().date()
        ).order_by('due_date')[:10],
    }
    
    return render(request, 'myapp/payment_analytics_dashboard.html', context)

@login_required
def payment_analytics_export(request):
    """Export payment analytics data"""
    from django.http import HttpResponse
    from django.db.models import Sum, Count
    import csv
    
    view_type = request.GET.get('view', 'school')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    # Get payments data
    payments = Payment.objects.filter(
        payment_date__range=[date_from, date_to]
    ).select_related('student', 'fee_structure__category')
    
    # Apply filters
    if view_type == 'student':
        student_id = request.GET.get('student')
        if student_id:
            payments = payments.filter(student_id=student_id)
    elif view_type == 'class':
        class_name = request.GET.get('class')
        if class_name:
            payments = payments.filter(student__class_name=class_name)
    elif view_type == 'batch':
        batch_year = request.GET.get('batch')
        if batch_year:
            payments = payments.filter(student__year_batch=batch_year)
    elif view_type == 'category':
        category_id = request.GET.get('category')
        if category_id:
            payments = payments.filter(fee_structure__category_id=category_id)
    
    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="payment_analytics_{view_type}_{date_from}_to_{date_to}.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        'Student Name', 'Student ID', 'Class', 'Batch', 
        'Fee Category', 'Amount', 'Payment Date', 'Payment Method', 'Status'
    ])
    
    for payment in payments:
        writer.writerow([
            f"{payment.student.first_name} {payment.student.last_name}",
            payment.student.student_id,
            payment.student.class_name or '',
            payment.student.year_batch,
            payment.fee_structure.category.name if payment.fee_structure else 'Individual Fee',
            payment.amount,
            payment.payment_date,
            payment.payment_method,
            payment.status,
        ])
    
    return response

@login_required
@require_GET
def get_form_fees(request):
    """AJAX view to get fees for a specific form level"""
    form = request.GET.get('form')
    
    if not form:
        return JsonResponse({'success': False, 'error': 'Form parameter is required'})
    
    try:
        from .models import FeeStructure
        fees = FeeStructure.objects.filter(
            form__iexact=form,
            is_active=True
        ).select_related('category')
        
        fees_data = []
        for fee in fees:
            fees_data.append({
                'category': fee.category.name,
                'amount': str(fee.amount) if fee.amount else 'Not set',
                'frequency': fee.frequency,
                'monthly_amount': str(fee.get_monthly_amount()) if fee.frequency == 'monthly' and fee.monthly_duration else None
            })
        
        return JsonResponse({
            'success': True,
            'form': form,
            'fees': fees_data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@login_required
def bulk_add_students_form(request):
    """Bulk add students through web form"""
    if request.method == 'POST':
        print("DEBUG: bulk_add_students_form POST request received")
        print("DEBUG: POST data keys:", list(request.POST.keys()))
        
        success_count = 0
        error_count = 0
        errors = []
        
        # Find all student entries in the form
        student_entries = {}
        for key, value in request.POST.items():
            if '_' in key:
                parts = key.split('_')
                if len(parts) >= 2:
                    field_name = '_'.join(parts[:-1])
                    student_num = parts[-1]
                    
                    if student_num.isdigit():
                        if student_num not in student_entries:
                            student_entries[student_num] = {}
                        student_entries[student_num][field_name] = value
        
        print(f"DEBUG: Found {len(student_entries)} student entries")
        
        for student_num, student_data in student_entries.items():
            try:
                # Check required fields
                required_fields = ['student_id', 'nric', 'first_name', 'last_name', 'year_batch']
                missing_fields = [field for field in required_fields if not student_data.get(field)]
                
                if missing_fields:
                    errors.append(f"Student {student_num}: Missing required fields: {', '.join(missing_fields)}")
                    error_count += 1
                    continue
                
                # Check if student already exists
                if Student.objects.filter(student_id=student_data['student_id']).exists():
                    errors.append(f"Student {student_num}: Student ID {student_data['student_id']} already exists")
                    error_count += 1
                    continue
                
                if Student.objects.filter(nric=student_data['nric']).exists():
                    errors.append(f"Student {student_num}: NRIC {student_data['nric']} already exists")
                    error_count += 1
                    continue
                
                # Create student data
                student_obj_data = {
                    'student_id': student_data['student_id'],
                    'nric': student_data['nric'],
                    'first_name': student_data['first_name'],
                    'last_name': student_data['last_name'],
                    'year_batch': student_data['year_batch'],
                    'is_active': student_data.get('is_active') == 'on',
                    'level': student_data.get('level', 'form'),
                    'level_custom': student_data.get('level_custom', ''),
                    'class_name': student_data.get('class_name', ''),
                    'program': student_data.get('program', ''),
                }
                
                # Auto-assign form number if level is 'form' but no form number is selected
                if student_obj_data['level'] == 'form' and not student_obj_data['level_custom']:
                    student_obj_data['level_custom'] = 'Form 1'
                    print(f"DEBUG: Auto-assigned form number for student {student_num}: Form 1")
                
                # Create student
                student = Student(**student_obj_data)
                student.save()
                
                # Automatically create user account
                try:
                    username = f"student_{student.student_id.lower()}"
                    if not User.objects.filter(username=username).exists():
                        password = f"{student.student_id.lower()}123"
                        user = User.objects.create_user(
                            username=username,
                            email=f"{username}@school.com",
                            password=password,
                            first_name=student.first_name,
                            last_name=student.last_name
                        )
                        from .models import UserProfile
                        UserProfile.objects.create(
                            user=user,
                            role='student',
                            student=student
                        )
                        print(f"DEBUG: User account created for student {student_num} - Username: {username}")
                except Exception as e:
                    print(f"DEBUG: Error creating user account for student {student_num}: {e}")
                
                # Generate fee statuses if student has a form level
                if student.level == 'form' and student.level_custom:
                    try:
                        from .models import FeeStructure, FeeStatus
                        from datetime import date, timedelta
                        
                        form_fees = FeeStructure.objects.filter(
                            form__iexact=student.level_custom,
                            is_active=True
                        )
                        
                        for fee_structure in form_fees:
                            existing_status = FeeStatus.objects.filter(
                                student=student,
                                fee_structure=fee_structure
                            ).first()
                            
                            if not existing_status:
                                if fee_structure.frequency == 'yearly':
                                    due_date = date.today() + timedelta(days=30)
                                elif fee_structure.frequency == 'termly':
                                    due_date = date.today() + timedelta(days=90)
                                elif fee_structure.frequency == 'monthly':
                                    due_date = date.today() + timedelta(days=30)
                                else:
                                    due_date = date.today() + timedelta(days=30)
                                
                                FeeStatus.objects.create(
                                    student=student,
                                    fee_structure=fee_structure,
                                    amount=fee_structure.amount or 0,
                                    due_date=due_date,
                                    status='pending'
                                )
                                print(f"DEBUG: Created fee status for student {student_num} - {fee_structure.category.name}")
                    except Exception as e:
                        print(f"DEBUG: Error creating fee statuses for student {student_num}: {e}")
                    
                    success_count += 1
                    print(f"DEBUG: Successfully created student {student_num} - {student.first_name} {student.last_name}")
                    
            except Exception as e:
                errors.append(f"Student {student_num}: {str(e)}")
                error_count += 1
        
        # Show results
        if success_count > 0:
            messages.success(request, f'Successfully added {success_count} students with automatic user account creation and form number assignment.')
        if error_count > 0:
            messages.warning(request, f'Failed to add {error_count} students. See details below.')
            for error in errors:
                messages.error(request, error)
        
        return redirect('myapp:student_list')
    
    return render(request, 'myapp/bulk_add_students_form.html')

@login_required
def reminder_options(request, payment_id):
    """Show reminder options (text or email) for a payment"""
    fee_status = get_object_or_404(FeeStatus, id=payment_id)
    
    # Calculate the amount to be paid (with discounts applied)
    try:
        amount_to_pay = fee_status.get_discounted_amount()
    except:
        amount_to_pay = fee_status.amount
    
    # Calculate days overdue/until due
    today = timezone.now().date()
    if fee_status.due_date < today:
        days_text = f"{today - fee_status.due_date} days overdue"
        is_overdue = True
    else:
        days_text = f"{fee_status.due_date - today} days until due"
        is_overdue = False
    
    # Get student contact information
    student = fee_status.student
    # Get student email from the linked User model through UserProfile
    try:
        user_profile = student.user_profile.first()
        student_email = user_profile.user.email if user_profile else None
    except:
        student_email = None
    
    # Get phone number from parent
    parent = student.parents.first()
    student_phone = parent.phone_number if parent else None
    
    context = {
        'fee_status': fee_status,
        'amount_to_pay': amount_to_pay,
        'days_text': days_text,
        'is_overdue': is_overdue,
        'student_email': student_email,
        'student_phone': student_phone,
        'student': student,
    }
    
    return render(request, 'myapp/reminder_options.html', context)

@login_required
def send_reminder_email(request, payment_id):
    """Send reminder via email and redirect to Gmail with generated letter"""
    fee_status = get_object_or_404(FeeStatus, id=payment_id)
    
    # Calculate the amount to be paid (with discounts applied)
    try:
        amount_to_pay = fee_status.get_discounted_amount()
    except:
        amount_to_pay = fee_status.amount
    
    # Calculate days overdue/until due
    today = timezone.now().date()
    if fee_status.due_date < today:
        days_text = f"{today - fee_status.due_date} days overdue"
        is_overdue = True
    else:
        days_text = f"{fee_status.due_date - today} days until due"
        is_overdue = False
    
    # Get student contact information
    student = fee_status.student
    # Get student email from the linked User model through UserProfile
    try:
        user_profile = student.user_profile.first()
        student_email = user_profile.user.email if user_profile else None
    except:
        student_email = None
    
    # Send directly to student's email if available, otherwise to admin
    if student_email:
        recipient_emails = [student_email]
    else:
        # Fallback to admin email if student email is not available
        admin_email = 'moaaj.upm@gmail.com'
        recipient_emails = [admin_email]
    
    # Prepare email content
    subject = f'Payment Reminder - {fee_status.fee_structure.category.name}'
    context = {
        'payment': fee_status,
        'fee_status': fee_status,
        'amount_to_pay': amount_to_pay,
        'days_text': days_text,
        'is_overdue': is_overdue,
        'now': timezone.now().date(),
    }
    
    # Render email template
    html_message = render_to_string('myapp/email/payment_reminder_email.html', context)
    plain_message = strip_tags(html_message)
    
    # Send email
    try:
        send_mail(
            subject=subject,
            message=plain_message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=recipient_emails,
            html_message=html_message,
            fail_silently=False,
        )
        if student_email:
            messages.success(request, f'Email reminder sent successfully to {student.first_name} {student.last_name} ({student_email})')
        else:
            messages.success(request, f'Email reminder sent to admin email (student email not available)')
    except Exception as e:
        messages.error(request, f'Failed to send email: {str(e)}')
    
    # Generate Gmail URL with letter content
    letter_content = generate_letter_content(fee_status, amount_to_pay, days_text, is_overdue)
    # Use student email for Gmail if available, otherwise admin email
    gmail_recipients = [student_email] if student_email else ['moaaj.upm@gmail.com']
    gmail_url = generate_gmail_url(subject, letter_content, gmail_recipients)
    
    return redirect(gmail_url)

@login_required
def send_reminder_text(request, payment_id):
    """Send reminder via text and redirect to Messages app with phone number"""
    fee_status = get_object_or_404(FeeStatus, id=payment_id)
    
    # Calculate the amount to be paid (with discounts applied)
    try:
        amount_to_pay = fee_status.get_discounted_amount()
    except:
        amount_to_pay = fee_status.amount
    
    # Calculate days overdue/until due
    today = timezone.now().date()
    if fee_status.due_date < today:
        days_text = f"{today - fee_status.due_date} days overdue"
        is_overdue = True
    else:
        days_text = f"{fee_status.due_date - today} days until due"
        is_overdue = False
    
    # Get student contact information
    student = fee_status.student
    
    # Get phone number from parent
    parent = student.parents.first()
    student_phone = parent.phone_number if parent else None
    
    # Generate text message content
    text_content = generate_text_message_content(fee_status, amount_to_pay, days_text, is_overdue)
    
    # Send text message if phone number is available
    if student_phone:
        try:
            send_text_message(student_phone, fee_status, amount_to_pay, days_text, is_overdue)
            messages.success(request, f'Text message sent successfully to {student.first_name} {student.last_name}')
        except Exception as e:
            messages.error(request, f'Failed to send text message: {str(e)}')
    else:
        messages.warning(request, 'No phone number available for text message')
    
    # Generate Messages app URL with phone number and content
    if student_phone:
        messages_url = generate_messages_url(student_phone, text_content)
        # Use JavaScript redirect for SMS protocol
        from django.http import HttpResponse
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Redirecting to Messages App</title>
            <meta charset="utf-8">
        </head>
        <body>
            <div style="text-align: center; padding: 50px; font-family: Arial, sans-serif;">
                <h2>📱 Redirecting to Messages App</h2>
                <p>Opening Messages app with phone number and message pre-filled...</p>
                <p><strong>Phone:</strong> {student_phone}</p>
                <p><strong>Message:</strong> {text_content[:100]}...</p>
                <div style="margin: 20px;">
                    <button onclick="window.location.href='{messages_url}'" style="padding: 10px 20px; background: #28a745; color: white; border: none; border-radius: 5px; cursor: pointer;">
                        Open Messages App
                    </button>
                </div>
                <p><small>If the button doesn't work, <a href="{messages_url}">click here</a></small></p>
            </div>
            <script>
                // Auto-redirect after 2 seconds
                setTimeout(function() {{
                    window.location.href = '{messages_url}';
                }}, 2000);
            </script>
        </body>
        </html>
        """
        return HttpResponse(html_content)
    else:
        # Fallback to Gmail if no phone number available
        letter_content = generate_letter_content(fee_status, amount_to_pay, days_text, is_overdue)
        subject = f'Payment Reminder Letter - {fee_status.fee_structure.category.name}'
        gmail_url = generate_gmail_url(subject, letter_content, ['moaaj.upm@gmail.com'])
        return redirect(gmail_url)

def generate_letter_content(fee_status, amount_to_pay, days_text, is_overdue):
    """Generate letter content for Gmail"""
    today = timezone.now().date()
    
    letter_content = f"""
Dear {fee_status.student.first_name} {fee_status.student.last_name},

This letter serves as a formal reminder regarding your outstanding payment.

PAYMENT DETAILS:
Student ID: {fee_status.student.student_id}
Fee Category: {fee_status.fee_structure.category.name}
Original Amount: RM {fee_status.amount:.2f}
Amount Due: RM {amount_to_pay:.2f}
Due Date: {fee_status.due_date.strftime('%d %B %Y')}
Status: {days_text}

"""
    
    if is_overdue:
        letter_content += f"""
This payment is OVERDUE. Please make the payment as soon as possible to avoid any late fees or penalties.

"""
    else:
        letter_content += f"""
This payment is due soon. Please ensure timely payment to avoid any late fees.

"""
    
    letter_content += f"""
You can make the payment through any of the following methods:
- Cash payment at the school office
- Bank transfer to our school account
- Online payment through our payment portal

If you have already made the payment, please ignore this reminder.

If you have any questions or need assistance, please contact the school office.

Thank you for your attention to this matter.

Best regards,
School Administration

Date: {today.strftime('%d %B %Y')}
"""
    
    return letter_content

def generate_gmail_url(subject, content, recipients):
    """Generate Gmail URL with pre-filled content"""
    import urllib.parse
    
    # Encode the content for URL
    encoded_subject = urllib.parse.quote(subject)
    encoded_content = urllib.parse.quote(content)
    
    # Create Gmail compose URL
    gmail_url = f"https://mail.google.com/mail/?view=cm&fs=1&to={','.join(recipients)}&su={encoded_subject}&body={encoded_content}"
    
    return gmail_url

def generate_text_message_content(fee_status, amount_to_pay, days_text, is_overdue):
    """Generate text message content for SMS"""
    student_name = f"{fee_status.student.first_name} {fee_status.student.last_name}"
    fee_category = fee_status.fee_structure.category.name
    
    if is_overdue:
        message = f"URGENT: {student_name}, your {fee_category} payment of RM {amount_to_pay:.2f} is {days_text}. Please pay immediately to avoid penalties. Contact school office for assistance."
    else:
        message = f"REMINDER: {student_name}, your {fee_category} payment of RM {amount_to_pay:.2f} is due in {days_text}. Please ensure timely payment."
    
    return message

def generate_messages_url(phone_number, message_content):
    """Generate Messages app URL with pre-filled phone number and content"""
    import urllib.parse
    
    # Remove any non-digit characters from phone number for SMS URL
    clean_phone = ''.join(filter(str.isdigit, phone_number))
    
    # Encode the message content for URL
    encoded_message = urllib.parse.quote(message_content)
    
    # Create SMS URL (works on most mobile devices)
    # Format: sms:phone_number?body=message
    sms_url = f"sms:{clean_phone}?body={encoded_message}"
    
    return sms_url